# -*- coding: utf-8 -*-
from flask import Flask, Response, request, jsonify, render_template, session, redirect, url_for
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from contextlib import contextmanager
import csv
import io
import json
import os
import secrets
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from threading import Lock

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

import psycopg2
import psycopg2.extras

# ── APP SETUP ─────────────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__, template_folder=BASE_DIR)

app.secret_key = os.environ.get('SECRET_KEY')
if not app.secret_key:
    if os.environ.get('FLASK_ENV') == 'production':
        raise RuntimeError(
            'SECRET_KEY environment variable must be set in production. '
            'Generate one with: python -c "import secrets; print(secrets.token_hex(32))"'
        )
    # Dev/local only -- never reaches production
    app.secret_key = 'mmsu_medical_dashboard_DEV_ONLY_not_for_production'
    print('[WARNING] SECRET_KEY not set -- using insecure dev default. Set SECRET_KEY env var.', file=sys.stderr)

app.config['SESSION_COOKIE_SECURE']   = os.environ.get('FLASK_ENV') == 'production'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)

ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin')

def _get_password_hash() -> str:
    """
    Load password hash from DB, falling back to env var / default.

    The DB value is validated before use: it must be a non-empty string that
    looks like a werkzeug hash (starts with 'pbkdf2:', 'scrypt:', or 'argon2:').
    A corrupted or unexpected value in app_settings is ignored and the env-var
    fallback is used instead, with a logged warning so ops can investigate.
    """
    try:
        with get_db() as conn:
            c = conn.cursor()
            c.execute("SELECT value FROM app_settings WHERE key='admin_password_hash'")
            row = c.fetchone()
            if row:
                value = row[0]
                if isinstance(value, str) and value.startswith(('pbkdf2:', 'scrypt:', 'argon2:')):
                    return value
                # Row exists but value looks wrong — log and fall through
                app.logger.warning(
                    '[auth] admin_password_hash in app_settings has unexpected format; '
                    'falling back to ADMIN_PASSWORD env var. Check the app_settings table.'
                )
    except Exception:
        pass  # DB not ready yet on first boot — fall through to env default
    _raw = os.environ.get('ADMIN_PASSWORD', 'mmsu2024')
    return generate_password_hash(_raw)

# ── DATABASE ──────────────────────────────────────────────────────────────────

def _get_db_url():
    url = os.environ.get('DATABASE_URL')
    if not url:
        raise RuntimeError('DATABASE_URL environment variable is not set.')
    # Heroku/Railway uses the legacy postgres:// scheme; psycopg2 needs postgresql://
    return url.replace('postgres://', 'postgresql://', 1)


# Connection pool -- reuses connections instead of opening a new one per request
from psycopg2 import pool as pg_pool

_pool: pg_pool.ThreadedConnectionPool | None = None
_pool_lock = Lock()

def get_pool() -> pg_pool.ThreadedConnectionPool:
    global _pool
    if _pool is None:
        with _pool_lock:
            if _pool is None:
                _pool = pg_pool.ThreadedConnectionPool(
                    minconn=2,
                    maxconn=10,
                    dsn=_get_db_url(),
                )
    return _pool


@contextmanager
def get_db():
    """Context manager that borrows a connection from the pool and always returns it."""
    pool = get_pool()
    conn = pool.getconn()
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        pool.putconn(conn)


def init_db():
    with get_db() as conn:
        c = conn.cursor()

        c.execute('''
            CREATE TABLE IF NOT EXISTS personnel (
                id         SERIAL PRIMARY KEY,
                name       TEXT,
                age        INTEGER,
                gender     TEXT,
                blood      TEXT,
                department TEXT,
                phone      TEXT,
                address    TEXT,
                conditions TEXT
            )
        ''')

        # Migrate older tables that predate the age/phone/address columns
        for col, coltype in [('age', 'INTEGER'), ('phone', 'TEXT'), ('address', 'TEXT')]:
            c.execute(f'''
                DO $$ BEGIN
                    ALTER TABLE personnel ADD COLUMN {col} {coltype};
                EXCEPTION WHEN duplicate_column THEN NULL;
                END $$;
            ''')

        c.execute('''
            CREATE TABLE IF NOT EXISTS visits (
                id           SERIAL PRIMARY KEY,
                personnel_id INTEGER REFERENCES personnel(id) ON DELETE CASCADE,
                visit_date   DATE NOT NULL,
                reason       TEXT,
                notes        TEXT,
                created_at   TIMESTAMP DEFAULT NOW()
            )
        ''')

        c.execute('''
            CREATE TABLE IF NOT EXISTS departments (
                id   SERIAL PRIMARY KEY,
                name TEXT UNIQUE NOT NULL
            )
        ''')

        c.execute('''
            CREATE TABLE IF NOT EXISTS audit_log (
                id         SERIAL PRIMARY KEY,
                username   TEXT,
                action     TEXT,
                detail     TEXT,
                ip         TEXT,
                created_at TIMESTAMP DEFAULT NOW()
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS app_settings (
                key   TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
        ''')

        c.execute('''
            CREATE TABLE IF NOT EXISTS notifications (
                id         SERIAL PRIMARY KEY,
                level      TEXT NOT NULL DEFAULT 'info',
                title      TEXT NOT NULL,
                body       TEXT,
                read       BOOLEAN NOT NULL DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT NOW()
            )
        ''')

        c.execute('''
            CREATE TABLE IF NOT EXISTS login_attempts (
                ip           TEXT NOT NULL,
                attempted_at TIMESTAMP NOT NULL DEFAULT NOW()
            )
        ''')
        c.execute('''
            CREATE INDEX IF NOT EXISTS idx_login_attempts_ip_time
            ON login_attempts (ip, attempted_at)
        ''')


# ── STARTUP HOOK ──────────────────────────────────────────────────────────────
# init_db() is intentionally NOT called at import time.
#
# Calling it at module level causes two problems:
#   1. Any DB unavailability at startup (e.g. Railway cold-start, bad DATABASE_URL)
#      crashes every gunicorn worker before a single request is handled, with no
#      useful HTTP error — just a silent process death.
#   2. With multiple workers (gunicorn -w 4), all workers import the module in
#      parallel and race to run the ALTER TABLE migrations simultaneously, which
#      can produce deadlocks or duplicate-column errors under load.
#
# Instead, init runs once per process on the first real request, guarded by a
# lock so only one thread does it even if several requests arrive concurrently.

_db_initialised = False
_init_lock       = Lock()


@app.before_request
def _ensure_db_ready():
    """Run init_db() exactly once per worker process, on the first request."""
    global _db_initialised
    if _db_initialised:          # fast path — no lock needed after first request
        return
    with _init_lock:
        if _db_initialised:      # re-check inside the lock (another thread may have won)
            return
        try:
            init_db()
            _db_initialised = True
        except Exception as e:
            # Surface the error as an HTTP 503 so ops can see it immediately
            # rather than getting cryptic 500s from every route handler.
            app.logger.exception('DB initialisation failed: %s', e)
            return (
                'Service temporarily unavailable — database initialisation failed. '
                'Check DATABASE_URL and server logs.',
                503,
            )


# ── HELPERS ───────────────────────────────────────────────────────────────────

def row_to_person(r):
    """
    Convert a personnel DB row to a plain dict.

    Accepts either a RealDictRow (from RealDictCursor) or a plain tuple so
    callers don't all have to be updated at once. New callers should use
    RealDictCursor so column access is by name, not position — that way a
    query reorder or extra column never silently corrupts the mapping.
    """
    return {
        'id':         r['id'],
        'name':       r['name'],
        'age':        r['age'],
        'gender':     r['gender'],
        'blood':      r['blood'],
        'department': r['department'],
        'phone':      r['phone'],
        'address':    r['address'],
        'conditions': r['conditions'].split('|') if r['conditions'] else [],
    }


def person_params(d):
    """Extract and normalise personnel fields from a request payload."""
    return (
        d.get('name', ''),
        d.get('age'),
        d.get('gender', ''),
        d.get('blood', ''),
        d.get('department', ''),
        d.get('phone', ''),
        d.get('address', ''),
        '|'.join(d.get('conditions', [])),
    )


def csv_response(rows, headers, filename):
    """Build a CSV download response from a list of rows."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return Response(
        buf.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


def audit(action, detail=''):
    """
    Write an entry to the audit log.

    On DB failure the error is logged via app.logger (captured by gunicorn /
    Railway) AND written to stderr as a plain-text fallback so the event is
    never silently dropped — it will appear in at least one log sink even if
    the structured logger is misconfigured.
    """
    try:
        with get_db() as conn:
            conn.cursor().execute(
                'INSERT INTO audit_log (username, action, detail, ip) VALUES (%s, %s, %s, %s)',
                (session.get('user', 'system'), action, detail, request.remote_addr),
            )
    except Exception:
        app.logger.error('[audit] Failed to write audit log (action=%s)', action, exc_info=True)
        # Stderr fallback: guarantees the event appears somewhere even if the
        # structured logger hasn't been set up yet (e.g. during DB init).
        print(
            f'[audit-fallback] action={action!r} detail={detail!r} '
            f'user={session.get("user", "system")!r} ip={request.remote_addr!r}',
            file=sys.stderr,
            flush=True,
        )


def notify(title, body='', level='info'):
    """Push an in-app notification (level: info | warning | danger)."""
    try:
        with get_db() as conn:
            conn.cursor().execute(
                'INSERT INTO notifications (level, title, body) VALUES (%s, %s, %s)',
                (level, title, body),
            )
    except Exception:
        app.logger.error('[notify] Failed to write notification (title=%s)', title, exc_info=True)

# ── INPUT VALIDATION ─────────────────────────────────────────────────────────

FIELD_LIMITS = {
    'name':       256,
    'gender':      16,
    'blood':        8,
    'department':  128,
    'phone':        32,
    'address':     512,
    'conditions': 1024,
}
VALID_GENDERS     = {'Male', 'Female', ''}
VALID_BLOOD_TYPES = {'A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-', ''}

# Single source of truth for high-risk conditions.
# Used by: search filters, add-personnel notifications, PDF export, all reports.
HIGH_RISK_CONDITIONS = {
    'Hypertension', 'Diabetes', 'Asthma',
    'Heart Disease', 'Tuberculosis', 'Cancer', 'HIV/AIDS', 'Epilepsy',
    'Chronic Kidney Disease', 'Stroke', 'Hepatitis B', 'Hepatitis C',
}


def validate_personnel(d: dict) -> str | None:
    """Return an error string if the payload is invalid, else None."""
    name = d.get('name', '').strip()
    if not name:
        return 'Name is required.'
    for field, limit in FIELD_LIMITS.items():
        val = d.get(field, '')
        if isinstance(val, str) and len(val) > limit:
            return f'Field "{field}" exceeds maximum length of {limit} characters.'
    age = d.get('age')
    if age is not None:
        try:
            age_int = int(age)
            if not (0 < age_int < 150):
                return 'Age must be between 1 and 149.'
        except (TypeError, ValueError):
            return 'Age must be a number.'
    if d.get('gender', '') not in VALID_GENDERS:
        return 'Invalid gender value.'
    if d.get('blood', '') not in VALID_BLOOD_TYPES:
        return 'Invalid blood type.'
    conditions = d.get('conditions', [])
    if not isinstance(conditions, list):
        return 'Conditions must be a list.'
    if len(conditions) > 30:
        return 'Too many conditions (max 30).'
    return None

# ── RATE LIMITER ──────────────────────────────────────────────────────────────
#
# Stored in PostgreSQL so all gunicorn workers share the same state.
# An in-memory dict would give each worker its own independent counter,
# letting an attacker make LOGIN_MAX_ATTEMPTS * num_workers attempts before
# being blocked.

LOGIN_MAX_ATTEMPTS = 10
LOGIN_WINDOW_SECS  = 300  # 5 minutes


def is_rate_limited(ip: str) -> bool:
    """
    Return True if this IP has hit the login attempt ceiling.

    Each call atomically:
      1. Prunes expired rows for this IP (keeps the table small).
      2. Counts recent attempts within the window.
      3. If under the limit, inserts a new attempt row and returns False.
      4. If at/over the limit, returns True without inserting.

    On DB error we fail open (return False) so a DB outage doesn't lock
    everyone out, but the error is logged for ops.
    """
    window_start = datetime.utcnow() - timedelta(seconds=LOGIN_WINDOW_SECS)
    try:
        with get_db() as conn:
            c = conn.cursor()
            # Prune stale rows for this IP first
            c.execute(
                'DELETE FROM login_attempts WHERE ip = %s AND attempted_at < %s',
                (ip, window_start),
            )
            c.execute(
                'SELECT COUNT(*) FROM login_attempts WHERE ip = %s AND attempted_at >= %s',
                (ip, window_start),
            )
            count = c.fetchone()[0]
            if count >= LOGIN_MAX_ATTEMPTS:
                return True
            c.execute('INSERT INTO login_attempts (ip) VALUES (%s)', (ip,))
            return False
    except Exception:
        app.logger.error('[rate_limit] DB error during rate limit check for %s', ip, exc_info=True)
        return False  # fail open — DB outage shouldn't lock out all users

# ── AUTH / CSRF DECORATORS ────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def csrf_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('X-CSRF-Token') or (request.json or {}).get('_csrf')
        if not token or token != session.get('csrf_token'):
            return jsonify({'error': 'Invalid CSRF token'}), 403
        return f(*args, **kwargs)
    return decorated


def get_csrf_token():
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)
    return session['csrf_token']

# ── AUTH ROUTES ───────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def index():
    return render_template('index.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        ip = request.remote_addr
        if is_rate_limited(ip):
            return jsonify({'success': False, 'error': 'Too many login attempts. Please wait 5 minutes.'}), 429
        data = request.json or {}
        if data.get('username') == ADMIN_USERNAME and check_password_hash(_get_password_hash(), data.get('password', '')):
            session.permanent = True
            session['user'] = data['username']
            audit('LOGIN', f'Admin logged in from {request.remote_addr}')
            return jsonify({'success': True})
        audit('LOGIN_FAIL', f'Failed login attempt for "{data.get("username")}"')
        notify('Failed login attempt', f'Username: {data.get("username","")} from {request.remote_addr}', 'warning')
        return jsonify({'success': False, 'error': 'Invalid credentials'}), 401
    return render_template('login.html')


@app.route('/logout', methods=['POST'])
@csrf_required
def logout():
    audit('LOGOUT')
    session.clear()
    return jsonify({'success': True})


@app.route('/csrf-token')
@login_required
def csrf_token():
    return jsonify({'token': get_csrf_token()})

# ── PERSONNEL ─────────────────────────────────────────────────────────────────

@app.route('/personnel')
@login_required
def get_personnel():
    """
    Returns all personnel (used by charts, reports, condition sidebar).
    For the paginated table use GET /personnel/search instead.
    """
    with get_db() as conn:
        c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        c.execute('SELECT id, name, age, gender, blood, department, phone, address, conditions FROM personnel ORDER BY name')
        return jsonify([row_to_person(r) for r in c.fetchall()])


@app.route('/personnel/search')
@login_required
def search_personnel():
    """
    Server-side filtered + paginated personnel list for the table view.

    Query params:
      q        – name/department text search (optional)
      dept     – exact department filter (optional)
      gender   – Male | Female (optional)
      blood    – blood type e.g. A+ (optional)
      risk     – 'high' | 'normal' (optional)
      page     – 1-based page number (default 1)
      per_page – rows per page (default 25, max 100)
    """
    q        = request.args.get('q', '').strip()
    dept     = request.args.get('dept', '').strip()
    gender   = request.args.get('gender', '').strip()
    blood    = request.args.get('blood', '').strip()
    risk     = request.args.get('risk', '').strip()   # 'high' | 'normal' | ''
    conds_raw = request.args.get('conditions', '').strip()   # comma-separated
    logic    = request.args.get('logic', 'OR').upper()       # 'AND' | 'OR'

    try:
        page     = max(1, int(request.args.get('page', 1)))
        per_page = min(100, max(1, int(request.args.get('per_page', 25))))
    except ValueError:
        return jsonify({'error': 'Invalid page or per_page parameter'}), 400

    # ── Build WHERE clause ────────────────────────────────────────────────────
    wheres, params = [], []

    if q:
        wheres.append('(LOWER(name) LIKE %s OR LOWER(department) LIKE %s)')
        like = f'%{q.lower()}%'
        params += [like, like]

    if dept:
        wheres.append('department = %s')
        params.append(dept)

    if gender:
        wheres.append('gender = %s')
        params.append(gender)

    if blood:
        wheres.append('blood = %s')
        params.append(blood)

    # ── conditions filter → SQL LIKE clauses ────────────────────────────────
    # The conditions column stores pipe-delimited values e.g. "Diabetes|Asthma".
    # We match each condition with %condition% which is safe: condition names
    # contain only letters/spaces and cannot partially match each other across
    # the pipe delimiter.
    if conds_raw:
        cond_list = [c2.strip() for c2 in conds_raw.split(',') if c2.strip()]
        if logic == 'AND':
            for cond in cond_list:
                wheres.append('conditions LIKE %s')
                params.append(f'%{cond}%')
        else:  # OR
            or_parts = ['conditions LIKE %s'] * len(cond_list)
            wheres.append('(' + ' OR '.join(or_parts) + ')')
            params.extend(f'%{cond}%' for cond in cond_list)

    # ── risk filter → SQL LIKE clauses ───────────────────────────────────────
    # Expand HIGH_RISK_CONDITIONS into OR-joined LIKE clauses so the DB engine
    # can use its own optimiser rather than loading every row into Python.
    if risk == 'high':
        risk_parts = ['conditions LIKE %s'] * len(HIGH_RISK_CONDITIONS)
        wheres.append('(' + ' OR '.join(risk_parts) + ')')
        params.extend(f'%{c2}%' for c2 in HIGH_RISK_CONDITIONS)
    elif risk == 'normal':
        for c2 in HIGH_RISK_CONDITIONS:
            wheres.append('conditions NOT LIKE %s')
            params.append(f'%{c2}%')

    where_sql = ('WHERE ' + ' AND '.join(wheres)) if wheres else ''

    with get_db() as conn:
        c = conn.cursor()

        c.execute(f'SELECT COUNT(*) FROM personnel {where_sql}', params)
        total = c.fetchone()[0]

        offset = (page - 1) * per_page
        c2 = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        c2.execute(
            f'SELECT id, name, age, gender, blood, department, phone, address, conditions '
            f'FROM personnel {where_sql} ORDER BY name LIMIT %s OFFSET %s',
            params + [per_page, offset],
        )
        page_rows = [row_to_person(r) for r in c2.fetchall()]

    return jsonify({
        'data':       page_rows,
        'total':      total,
        'page':       page,
        'per_page':   per_page,
        'total_pages': max(1, -(-total // per_page)),  # ceiling division
    })


@app.route('/personnel/add', methods=['POST'])
@login_required
@csrf_required
def add_personnel():
    d = request.json or {}
    err = validate_personnel(d)
    if err:
        return jsonify({'error': err}), 400
    with get_db() as conn:
        conn.cursor().execute(
            'INSERT INTO personnel (name, age, gender, blood, department, phone, address, conditions) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)',
            person_params(d),
        )
    audit('ADD_PERSONNEL', d.get('name', ''))
    hr_conds = [c for c in d.get('conditions', []) if c in HIGH_RISK_CONDITIONS]
    if hr_conds:
        notify(
            f'High-risk personnel added',
            f'{d.get("name","")} — {", ".join(hr_conds)}',
            'danger',
        )
    return jsonify({'message': 'Personnel added successfully!'})


@app.route('/personnel/update/<int:pid>', methods=['PUT'])
@login_required
@csrf_required
def update_personnel(pid):
    d = request.json or {}
    err = validate_personnel(d)
    if err:
        return jsonify({'error': err}), 400
    with get_db() as conn:
        c = conn.cursor()
        c.execute('SELECT id FROM personnel WHERE id = %s', (pid,))
        if not c.fetchone():
            return jsonify({'error': 'Not found'}), 404
        c.execute(
            '''UPDATE personnel
               SET name=%s, age=%s, gender=%s, blood=%s,
                   department=%s, phone=%s, address=%s, conditions=%s
               WHERE id=%s''',
            (*person_params(d), pid),
        )
    audit('UPDATE_PERSONNEL', f'id={pid} name={d.get("name", "")}')
    return jsonify({'message': 'Updated successfully!'})


@app.route('/personnel/delete/<int:pid>', methods=['DELETE'])
@login_required
@csrf_required
def delete_personnel(pid):
    with get_db() as conn:
        c = conn.cursor()
        c.execute('SELECT name FROM personnel WHERE id = %s', (pid,))
        row = c.fetchone()
        if not row:
            return jsonify({'error': 'Not found'}), 404
        c.execute('DELETE FROM personnel WHERE id = %s', (pid,))
    audit('DELETE_PERSONNEL', f'id={pid} name={row[0]}')
    return jsonify({'message': 'Deleted successfully!'})


def _parse_csv(content: str) -> tuple[list, list[str]]:
    """
    Parse and validate CSV content.

    Returns (records, errors) where:
      - records is a list of tuples ready for DB insert
      - errors is a list of human-readable error strings (empty on success)

    Validation is exhaustive: ALL row errors are collected before returning,
    so the user sees every problem at once instead of fixing one at a time.
    No database is touched.
    """
    reader = csv.DictReader(io.StringIO(content))
    records = []
    errors  = []

    for i, row in enumerate(reader, start=2):
        raw_conditions = row.get('conditions', '')
        cond_list = [c.strip() for c in raw_conditions.split('|') if c.strip()] if raw_conditions else []
        d = {
            'name':       row.get('name', '').strip(),
            'age':        row.get('age') or None,
            'gender':     row.get('gender', '').strip(),
            'blood':      row.get('blood', '').strip(),
            'department': row.get('department', '').strip(),
            'phone':      row.get('phone', '').strip(),
            'address':    row.get('address', '').strip(),
            'conditions': cond_list,
        }
        err = validate_personnel(d)
        if err:
            errors.append(f'Row {i}: {err}')
        else:
            records.append((
                d['name'], d['age'], d['gender'], d['blood'],
                d['department'], d['phone'], d['address'], raw_conditions,
            ))

    return records, errors


def _snapshot_personnel(conn) -> list[dict]:
    """Return all current personnel rows as a plain list of dicts (for pre-import backup)."""
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    c.execute('SELECT id, name, age, gender, blood, department, phone, address, conditions FROM personnel ORDER BY id')
    return [dict(r) for r in c.fetchall()]


@app.route('/upload', methods=['POST'])
@login_required
@csrf_required
def upload():
    """
    CSV import endpoint.

    Query params:
      mode=preview   – validate only; returns a summary without touching the DB.
                       Safe to call as a dry-run before committing.
      (default)      – validate then replace all personnel in a single transaction.

    The existing personnel table is snapshotted before the replace and returned
    in the response as `backup` so the caller can offer an undo/download.

    The DELETE + INSERT runs inside one transaction: if the insert fails for any
    reason the delete is rolled back and the original data is preserved.
    """
    preview_only = request.args.get('mode') == 'preview'

    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'No file uploaded'}), 400

    try:
        content = file.read().decode('utf-8')
    except UnicodeDecodeError:
        return jsonify({'error': 'File must be UTF-8 encoded'}), 400

    records, errors = _parse_csv(content)

    # ── Always return all validation errors at once ───────────────────────────
    if errors:
        return jsonify({
            'error': f'{len(errors)} validation error(s) found. Fix them all before re-uploading.',
            'validation_errors': errors,
        }), 400

    if not records:
        return jsonify({'error': 'CSV file contains no data rows'}), 400

    # ── Preview mode: return what would happen without changing anything ──────
    if preview_only:
        with get_db() as conn:
            existing_count = conn.cursor()
            existing_count.execute('SELECT COUNT(*) FROM personnel')
            current_count = existing_count.fetchone()[0]
        return jsonify({
            'preview': True,
            'incoming_records': len(records),
            'current_records':  current_count,
            'message': (
                f'CSV is valid. Importing will replace {current_count} existing '
                f'record(s) with {len(records)} new record(s).'
            ),
        })

    # ── Commit: snapshot → DELETE → INSERT in one transaction ────────────────
    with get_db() as conn:
        # Snapshot existing data BEFORE the delete so the caller can offer
        # a rollback/download even after the import succeeds.
        snapshot = _snapshot_personnel(conn)

        c = conn.cursor()
        c.execute('DELETE FROM personnel')
        psycopg2.extras.execute_batch(
            c,
            'INSERT INTO personnel (name, age, gender, blood, department, phone, address, conditions)'
            ' VALUES (%s,%s,%s,%s,%s,%s,%s,%s)',
            records,
        )
        # Keep the serial sequence in sync with the new rows.
        c.execute("SELECT setval('personnel_id_seq', (SELECT COALESCE(MAX(id), 0) FROM personnel))")
        # conn.commit() is called automatically by the get_db() context manager.

    audit('UPLOAD_CSV', f'{len(records)} records imported, {len(snapshot)} replaced')
    notify(
        'CSV Import complete',
        f'{len(records)} personnel records imported (replaced {len(snapshot)}).',
        'info',
    )
    return jsonify({
        'message':          f'{len(records)} records uploaded successfully!',
        'imported':         len(records),
        'replaced':         len(snapshot),
        'backup':           snapshot,          # caller may offer this as a downloadable undo
        'backup_timestamp': datetime.now().isoformat(),
    })

# ── VISITS ────────────────────────────────────────────────────────────────────

@app.route('/personnel/<int:pid>/visits')
@login_required
def get_visits(pid):
    with get_db() as conn:
        c = conn.cursor()
        c.execute(
            'SELECT id, visit_date, reason, notes, created_at FROM visits WHERE personnel_id = %s ORDER BY visit_date DESC',
            (pid,),
        )
        return jsonify([
            {'id': r[0], 'visit_date': str(r[1]), 'reason': r[2], 'notes': r[3], 'created_at': str(r[4])}
            for r in c.fetchall()
        ])


@app.route('/personnel/<int:pid>/visits/add', methods=['POST'])
@login_required
@csrf_required
def add_visit(pid):
    d = request.json or {}
    with get_db() as conn:
        c = conn.cursor()
        c.execute('SELECT id FROM personnel WHERE id = %s', (pid,))
        if not c.fetchone():
            return jsonify({'error': 'Personnel not found'}), 404
        c.execute(
            'INSERT INTO visits (personnel_id, visit_date, reason, notes) VALUES (%s, %s, %s, %s)',
            (pid, d.get('visit_date'), d.get('reason', ''), d.get('notes', '')),
        )
    audit('ADD_VISIT', f'personnel_id={pid}')
    return jsonify({'message': 'Visit recorded!'})


@app.route('/visits/delete/<int:vid>', methods=['DELETE'])
@login_required
@csrf_required
def delete_visit(vid):
    with get_db() as conn:
        c = conn.cursor()
        c.execute('SELECT id FROM visits WHERE id = %s', (vid,))
        if not c.fetchone():
            return jsonify({'error': 'Visit not found'}), 404
        c.execute('DELETE FROM visits WHERE id = %s', (vid,))
    audit('DELETE_VISIT', f'visit_id={vid}')
    return jsonify({'message': 'Visit deleted!'})

# ── DEPARTMENTS ───────────────────────────────────────────────────────────────

@app.route('/departments')
@login_required
def get_departments():
    with get_db() as conn:
        c = conn.cursor()
        c.execute('SELECT id, name FROM departments ORDER BY name')
        return jsonify([{'id': r[0], 'name': r[1]} for r in c.fetchall()])


@app.route('/departments/add', methods=['POST'])
@login_required
@csrf_required
def add_department():
    name = (request.json or {}).get('name', '').strip()
    if not name:
        return jsonify({'error': 'Name required'}), 400
    try:
        with get_db() as conn:
            conn.cursor().execute('INSERT INTO departments (name) VALUES (%s)', (name,))
    except psycopg2.errors.UniqueViolation:
        return jsonify({'error': 'Department already exists'}), 409
    audit('ADD_DEPT', name)
    return jsonify({'message': f'Department "{name}" added!'})


@app.route('/departments/delete/<int:did>', methods=['DELETE'])
@login_required
@csrf_required
def delete_department(did):
    with get_db() as conn:
        c = conn.cursor()
        c.execute('SELECT name FROM departments WHERE id = %s', (did,))
        row = c.fetchone()
        if not row:
            return jsonify({'error': 'Not found'}), 404
        c.execute('DELETE FROM departments WHERE id = %s', (did,))
    audit('DELETE_DEPT', row[0])
    return jsonify({'message': 'Department deleted!'})

# ── AUDIT LOG ─────────────────────────────────────────────────────────────────

@app.route('/audit-log')
@login_required
def get_audit_log():
    """
    Filterable, paginated audit log.

    Query params:
      q        – free-text search across action + detail (optional)
      action   – exact action filter e.g. LOGIN, DELETE_PERSONNEL (optional)
      date_from – YYYY-MM-DD inclusive start date (optional)
      date_to   – YYYY-MM-DD inclusive end date (optional)
      page     – 1-based page (default 1)
      per_page – rows per page (default 50, max 200)
    """
    q         = request.args.get('q', '').strip()
    action_f  = request.args.get('action', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to   = request.args.get('date_to', '').strip()

    try:
        page     = max(1, int(request.args.get('page', 1)))
        per_page = min(200, max(1, int(request.args.get('per_page', 50))))
    except ValueError:
        return jsonify({'error': 'Invalid page or per_page'}), 400

    wheres, params = [], []

    if q:
        wheres.append('(LOWER(action) LIKE %s OR LOWER(detail) LIKE %s OR LOWER(COALESCE(ip,\'\')) LIKE %s)')
        like = f'%{q.lower()}%'
        params += [like, like, like]

    if action_f:
        wheres.append('action = %s')
        params.append(action_f)

    if date_from:
        wheres.append('created_at >= %s')
        params.append(date_from)

    if date_to:
        wheres.append('created_at < (%s::date + INTERVAL \'1 day\')')
        params.append(date_to)

    where_sql = ('WHERE ' + ' AND '.join(wheres)) if wheres else ''

    with get_db() as conn:
        c = conn.cursor()
        c.execute(f'SELECT COUNT(*) FROM audit_log {where_sql}', params)
        total = c.fetchone()[0]

        offset = (page - 1) * per_page
        c.execute(
            f'SELECT id, username, action, detail, ip, created_at FROM audit_log {where_sql} '
            f'ORDER BY created_at DESC LIMIT %s OFFSET %s',
            params + [per_page, offset],
        )
        rows = [
            {'id': r[0], 'username': r[1], 'action': r[2], 'detail': r[3], 'ip': r[4], 'created_at': str(r[5])}
            for r in c.fetchall()
        ]

    return jsonify({
        'data':        rows,
        'total':       total,
        'page':        page,
        'per_page':    per_page,
        'total_pages': max(1, -(-total // per_page)),
    })


@app.route('/audit-log/actions')
@login_required
def get_audit_actions():
    """Return sorted list of distinct action types for the filter dropdown."""
    with get_db() as conn:
        c = conn.cursor()
        c.execute('SELECT DISTINCT action FROM audit_log ORDER BY action')
        return jsonify([r[0] for r in c.fetchall()])


@app.route('/audit-log/export')
@login_required
def export_audit_log():
    """Download the full audit log (optionally filtered) as CSV."""
    q         = request.args.get('q', '').strip()
    action_f  = request.args.get('action', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to   = request.args.get('date_to', '').strip()

    wheres, params = [], []
    if q:
        wheres.append('(LOWER(action) LIKE %s OR LOWER(detail) LIKE %s OR LOWER(COALESCE(ip,\'\')) LIKE %s)')
        like = f'%{q.lower()}%'
        params += [like, like, like]
    if action_f:
        wheres.append('action = %s')
        params.append(action_f)
    if date_from:
        wheres.append('created_at >= %s')
        params.append(date_from)
    if date_to:
        wheres.append('created_at < (%s::date + INTERVAL \'1 day\')')
        params.append(date_to)

    where_sql = ('WHERE ' + ' AND '.join(wheres)) if wheres else ''

    with get_db() as conn:
        c = conn.cursor()
        c.execute(
            f'SELECT id, username, action, detail, ip, created_at FROM audit_log {where_sql} ORDER BY created_at DESC',
            params,
        )
        rows = c.fetchall()

    audit('EXPORT_AUDIT_LOG', f'{len(rows)} entries')
    return csv_response(
        rows,
        ['id', 'username', 'action', 'detail', 'ip', 'created_at'],
        'audit_log_export.csv',
    )

# ── EXPORT ────────────────────────────────────────────────────────────────────

@app.route('/export/personnel')
@login_required
def export_personnel():
    with get_db() as conn:
        c = conn.cursor()
        c.execute('SELECT id, name, age, gender, blood, department, phone, address, conditions FROM personnel ORDER BY id')
        rows = c.fetchall()
    audit('EXPORT_CSV', f'{len(rows)} records')
    return csv_response(rows, ['id', 'name', 'age', 'gender', 'blood', 'department', 'phone', 'address', 'conditions'], 'personnel_export.csv')


@app.route('/export/visits')
@login_required
def export_visits():
    with get_db() as conn:
        c = conn.cursor()
        c.execute('''
            SELECT v.id, p.name, v.visit_date, v.reason, v.notes
            FROM visits v
            JOIN personnel p ON p.id = v.personnel_id
            ORDER BY v.visit_date DESC
        ''')
        rows = c.fetchall()
    audit('EXPORT_VISITS_CSV', f'{len(rows)} visits')
    return csv_response(rows, ['visit_id', 'personnel_name', 'visit_date', 'reason', 'notes'], 'visits_export.csv')

# ── PDF EXPORT ────────────────────────────────────────────────────────────────

@app.route('/personnel/<int:pid>/pdf')
@login_required
def export_personnel_pdf(pid):
    if not REPORTLAB_AVAILABLE:
        return jsonify({'error': 'reportlab is not installed on the server. Run: pip install reportlab'}), 500

    with get_db() as conn:
        c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        c.execute('SELECT id, name, age, gender, blood, department, phone, address, conditions FROM personnel WHERE id = %s', (pid,))
        row = c.fetchone()
        if not row:
            return jsonify({'error': 'Personnel not found'}), 404
        p = row_to_person(row)

        c2 = conn.cursor()
        c2.execute(
            'SELECT visit_date, reason, notes FROM visits WHERE personnel_id = %s ORDER BY visit_date DESC LIMIT 10',
            (pid,)
        )
        visits = c2.fetchall()

    # Optional AI suggestions & risk prediction passed from the frontend as URL-encoded JSON
    import urllib.parse
    ai_suggestions = []
    ai_interaction = ''
    ai_risk = None

    ai_json = request.args.get('ai_data', '')
    if ai_json:
        try:
            ai_payload = json.loads(urllib.parse.unquote(ai_json))
            ai_suggestions = ai_payload.get('suggestions', [])
            ai_interaction = ai_payload.get('interaction_note', '')
        except (json.JSONDecodeError, ValueError):
            pass

    risk_json = request.args.get('ai_risk', '')
    if risk_json:
        try:
            ai_risk = json.loads(urllib.parse.unquote(risk_json))
        except (json.JSONDecodeError, ValueError):
            pass

    audit('EXPORT_PDF', f'id={pid} name={p["name"]}')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    GREEN       = colors.HexColor('#1a7a3c')
    GREEN_LIGHT = colors.HexColor('#e8f5ee')
    GOLD        = colors.HexColor('#d4b84a')
    DARK        = colors.HexColor('#111111')
    GREY        = colors.HexColor('#555555')
    LIGHT_GREY  = colors.HexColor('#f6f7f6')
    BORDER      = colors.HexColor('#e0e0e0')
    RED_BG      = colors.HexColor('#fdecea')
    RED_TEXT    = colors.HexColor('#c0392b')

    styles = getSampleStyleSheet()

    def style(name, **kw):
        return ParagraphStyle(name, parent=styles['Normal'], **kw)

    title_style    = style('Title',    fontSize=20, textColor=DARK,  fontName='Helvetica-Bold', spaceAfter=2)
    dept_style     = style('Dept',     fontSize=11, textColor=GREY,  fontName='Helvetica')
    label_style    = style('Label',    fontSize=8,  textColor=GREY,  fontName='Helvetica-Bold')
    value_style    = style('Value',    fontSize=12, textColor=DARK,  fontName='Helvetica-Bold')
    value_sm       = style('ValueSm',  fontSize=11, textColor=DARK,  fontName='Helvetica')
    section_style  = style('Section',  fontSize=8,  textColor=GREY,  fontName='Helvetica-Bold', spaceBefore=14, spaceAfter=6)
    footer_style   = style('Footer',   fontSize=8,  textColor=GREY,  fontName='Helvetica', alignment=TA_CENTER)
    visit_date_s   = style('VDate',    fontSize=8,  textColor=GREEN, fontName='Helvetica-Bold')
    visit_reason_s = style('VReason',  fontSize=11, textColor=DARK,  fontName='Helvetica-Bold')
    visit_notes_s  = style('VNotes',   fontSize=10, textColor=GREY,  fontName='Helvetica')

    initials = ''.join(w[0] for w in p['name'].split() if w)[:2].upper()

    story = []
    page_w = A4[0] - 4*cm

    name_dept_combined = style('NameDept', fontSize=11, textColor=GREY, fontName='Helvetica', leading=28, leftIndent=10)
    name_para = Paragraph(
        f'<font name="Helvetica-Bold" size="20" color="#111111">{p["name"]}</font><br/>'
        f'<font name="Helvetica" size="11" color="#555555">{p["department"]} Department</font>',
        name_dept_combined
    )

    logo_combined = style('LogoCombined', fontSize=9, textColor=GREY, fontName='Helvetica', alignment=TA_RIGHT, leading=20)
    logo_para = Paragraph(
        f'<font name="Helvetica-Bold" size="13" color="#1a7a3c">MMSU Medical</font><br/>'
        f'<font name="Helvetica" size="9" color="#555555">Health Records System</font>',
        logo_combined
    )

    header_data = [[
        Paragraph(f'<font color="#d4b84a"><b>{initials}</b></font>',
                  style('Av', fontSize=22, fontName='Helvetica-Bold', alignment=TA_CENTER, textColor=GOLD)),
        name_para,
        logo_para,
    ]]
    header_table = Table(header_data, colWidths=[2*cm, 10*cm, 5*cm])
    header_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(0,0), GREEN),
        ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
        ('ALIGN',         (0,0),(0,0), 'CENTER'),
        ('ALIGN',         (2,0),(2,0), 'RIGHT'),
        ('TOPPADDING',    (0,0),(-1,-1), 14),
        ('BOTTOMPADDING', (0,0),(-1,-1), 14),
        ('LEFTPADDING',   (0,0),(0,0), 0),
        ('RIGHTPADDING',  (0,0),(0,0), 0),
        ('LEFTPADDING',   (2,0),(2,0), 0),
        ('RIGHTPADDING',  (2,0),(2,0), 0),
    ]))
    story.append(header_table)
    story.append(HRFlowable(width='100%', thickness=2, color=GREEN, spaceAfter=12, spaceBefore=10))

    def info_cell(label, value, blood=False):
        val_color = GOLD if blood else DARK
        return [
            Paragraph(label, label_style),
            Paragraph(str(value) if value else '—', style('VC', fontSize=12, textColor=val_color, fontName='Helvetica-Bold')),
        ]

    age_val   = f'{p["age"]} yrs' if p.get('age') else '—'
    grid_data = [
        [info_cell('GENDER', p.get('gender') or '—'),  info_cell('BLOOD TYPE', p.get('blood') or '—', blood=True)],
        [info_cell('AGE',    age_val),                  info_cell('PHONE', p.get('phone') or '—')],
    ]

    def make_cell(items):
        tbl = Table([[i] for i in items], colWidths=[(page_w/2 - 0.3*cm)])
        tbl.setStyle(TableStyle([
            ('BACKGROUND',     (0,0),(-1,-1), LIGHT_GREY),
            ('BOX',            (0,0),(-1,-1), 0.5, BORDER),
            ('TOPPADDING',     (0,0),(-1,-1), 8),
            ('BOTTOMPADDING',  (0,0),(-1,-1), 8),
            ('LEFTPADDING',    (0,0),(-1,-1), 12),
            ('RIGHTPADDING',   (0,0),(-1,-1), 12),
        ]))
        return tbl

    grid_table_data = [
        [make_cell(grid_data[0][0]), make_cell(grid_data[0][1])],
        [make_cell(grid_data[1][0]), make_cell(grid_data[1][1])],
    ]
    grid_table = Table(grid_table_data, colWidths=[page_w/2 - 0.15*cm, page_w/2 - 0.15*cm], hAlign='LEFT')
    grid_table.setStyle(TableStyle([
        ('VALIGN',       (0,0),(-1,-1), 'TOP'),
        ('LEFTPADDING',  (0,0),(-1,-1), 0),
        ('RIGHTPADDING', (0,0),(-1,-1), 0),
        ('TOPPADDING',   (0,0),(-1,-1), 3),
        ('BOTTOMPADDING',(0,0),(-1,-1), 3),
        ('INNERGRID',    (0,0),(-1,-1), 0, colors.white),
        ('BOX',          (0,0),(-1,-1), 0, colors.white),
    ]))
    story.append(grid_table)
    story.append(Spacer(1, 6))

    addr_cell = Table([
        [Paragraph('ADDRESS', label_style)],
        [Paragraph(p.get('address') or '—', value_sm)],
    ], colWidths=[page_w])
    addr_cell.setStyle(TableStyle([
        ('BACKGROUND',   (0,0),(-1,-1), LIGHT_GREY),
        ('BOX',          (0,0),(-1,-1), 0.5, BORDER),
        ('TOPPADDING',   (0,0),(-1,-1), 8),
        ('BOTTOMPADDING',(0,0),(-1,-1), 8),
        ('LEFTPADDING',  (0,0),(-1,-1), 12),
        ('RIGHTPADDING', (0,0),(-1,-1), 12),
    ]))
    story.append(addr_cell)

    story.append(Paragraph('RECORDED CONDITIONS', section_style))
    if p['conditions']:
        cond_cells = []
        for cond in p['conditions']:
            is_high = cond in HIGH_RISK_CONDITIONS
            bg      = RED_BG   if is_high else LIGHT_GREY
            txt     = RED_TEXT if is_high else DARK
            border  = colors.HexColor('#f5c6c6') if is_high else BORDER
            cell = Table([[Paragraph(cond, style('Pill', fontSize=10, textColor=txt, fontName='Helvetica-Bold'))]],
                         colWidths=[len(cond)*6 + 24])
            cell.setStyle(TableStyle([
                ('BACKGROUND',   (0,0),(-1,-1), bg),
                ('BOX',          (0,0),(-1,-1), 0.5, border),
                ('TOPPADDING',   (0,0),(-1,-1), 4),
                ('BOTTOMPADDING',(0,0),(-1,-1), 4),
                ('LEFTPADDING',  (0,0),(-1,-1), 10),
                ('RIGHTPADDING', (0,0),(-1,-1), 10),
            ]))
            cond_cells.append(cell)

        MAX_ROW_W = page_w
        rows, row_cells, row_widths, row_w = [], [], [], 0
        for cell, cond in zip(cond_cells, p['conditions']):
            cell_w = len(cond) * 6 + 24
            if row_cells and row_w + cell_w > MAX_ROW_W:
                rows.append((row_cells, row_widths))
                row_cells, row_widths, row_w = [], [], 0
            row_cells.append(cell)
            row_widths.append(cell_w)
            row_w += cell_w + 4
        if row_cells:
            rows.append((row_cells, row_widths))

        for r_cells, r_widths in rows:
            pills_table = Table([r_cells], colWidths=r_widths, hAlign='LEFT')
            pills_table.setStyle(TableStyle([
                ('LEFTPADDING',   (0,0),(-1,-1), 0),
                ('RIGHTPADDING',  (0,0),(-1,-1), 4),
                ('TOPPADDING',    (0,0),(-1,-1), 0),
                ('BOTTOMPADDING', (0,0),(-1,-1), 4),
            ]))
            story.append(pills_table)
    else:
        story.append(Paragraph('No conditions recorded.', style('NoCond', fontSize=11, textColor=GREY, fontName='Helvetica')))

    if visits:
        story.append(Paragraph('RECENT VISITS', section_style))
        for vdate, reason, notes in visits:
            visit_data = [
                [Paragraph(str(vdate), visit_date_s)],
                [Paragraph(reason or 'General Visit', visit_reason_s)],
            ]
            if notes:
                visit_data.append([Paragraph(notes, visit_notes_s)])
            vt = Table(visit_data, colWidths=[page_w - 0.3*cm])
            vt.setStyle(TableStyle([
                ('BACKGROUND',   (0,0),(-1,-1), LIGHT_GREY),
                ('BOX',          (0,0),(-1,-1), 0.5, BORDER),
                ('TOPPADDING',   (0,0),(-1,-1), 6),
                ('BOTTOMPADDING',(0,0),(-1,-1), 6),
                ('LEFTPADDING',  (0,0),(-1,-1), 12),
                ('RIGHTPADDING', (0,0),(-1,-1), 12),
            ]))
            story.append(vt)
            story.append(Spacer(1, 5))

    # ── AI TREATMENT SUGGESTIONS (optional) ──
    if ai_suggestions:
        story.append(Paragraph('AI TREATMENT SUGGESTIONS', section_style))
        story.append(Paragraph(
            'AI suggestions are for general reference only. Always defer to a licensed physician for final treatment decisions.',
            style('AiDisclaimer', fontSize=8, textColor=GREY, fontName='Helvetica', spaceAfter=6)
        ))
        for s in ai_suggestions:
            ai_data_rows = [
                [Paragraph(s.get('condition', ''), style('AiCond', fontSize=8, textColor=GREEN, fontName='Helvetica-Bold'))],
                [Paragraph(s.get('medicine', ''), style('AiMed', fontSize=11, textColor=DARK, fontName='Helvetica-Bold'))],
                [Paragraph(s.get('description', ''), style('AiDesc', fontSize=10, textColor=GREY, fontName='Helvetica'))],
            ]
            if s.get('warning'):
                ai_data_rows.append([Paragraph(f'\u26a0 {s["warning"]}', style('AiWarn', fontSize=10, textColor=RED_TEXT, fontName='Helvetica'))])
            ai_tbl = Table(ai_data_rows, colWidths=[page_w - 0.3*cm])
            ai_tbl.setStyle(TableStyle([
                ('BACKGROUND',    (0,0),(-1,-1), GREEN_LIGHT),
                ('BOX',           (0,0),(-1,-1), 0.5, colors.HexColor('#b2d8c0')),
                ('TOPPADDING',    (0,0),(-1,-1), 6),
                ('BOTTOMPADDING', (0,0),(-1,-1), 6),
                ('LEFTPADDING',   (0,0),(-1,-1), 12),
                ('RIGHTPADDING',  (0,0),(-1,-1), 12),
            ]))
            story.append(ai_tbl)
            story.append(Spacer(1, 5))
        if ai_interaction:
            story.append(Paragraph(
                f'\u26a0 Interaction Note: {ai_interaction}',
                style('AiInteract', fontSize=10, textColor=RED_TEXT, fontName='Helvetica',
                      backColor=RED_BG, spaceAfter=6, leftIndent=8, rightIndent=8)
            ))

    # ── AI RISK PREDICTION (optional) ──
    if ai_risk:
        story.append(Paragraph('AI RISK PREDICTION', section_style))
        story.append(Paragraph(
            'Risk prediction is AI-generated for reference only. Consult a licensed physician for clinical decisions.',
            style('RiskDisclaimer', fontSize=8, textColor=GREY, fontName='Helvetica', spaceAfter=6)
        ))

        risk_score  = ai_risk.get('risk_score', 0)
        risk_level  = ai_risk.get('risk_level', '')
        risk_color  = ai_risk.get('risk_color', 'warn')
        risk_colors_map = {
            'safe':     colors.HexColor('#4ec97a'),
            'warn':     colors.HexColor('#e8c84a'),
            'danger':   colors.HexColor('#d95555'),
            'critical': colors.HexColor('#b42828'),
        }
        rc = risk_colors_map.get(risk_color, risk_colors_map['warn'])

        # Score + level row
        score_tbl = Table([[
            Paragraph(f'<font size="26" color="{rc.hexval()}">{risk_score}</font>',
                      style('RScore', fontSize=26, fontName='Helvetica-Bold', alignment=TA_CENTER)),
            Paragraph(
                f'<font name="Helvetica-Bold" size="13">{risk_level} Risk</font><br/>'
                f'<font name="Helvetica" size="9" color="#555555">{ai_risk.get("future_outlook", "")}</font>',
                style('RLevel', fontSize=11, fontName='Helvetica', leading=16)
            ),
        ]], colWidths=[2.5*cm, page_w - 2.5*cm])
        score_tbl.setStyle(TableStyle([
            ('BACKGROUND',   (0,0),(-1,-1), LIGHT_GREY),
            ('BOX',          (0,0),(-1,-1), 0.5, BORDER),
            ('VALIGN',       (0,0),(-1,-1), 'MIDDLE'),
            ('ALIGN',        (0,0),(0,0), 'CENTER'),
            ('TOPPADDING',   (0,0),(-1,-1), 10),
            ('BOTTOMPADDING',(0,0),(-1,-1), 10),
            ('LEFTPADDING',  (0,0),(-1,-1), 12),
            ('RIGHTPADDING', (0,0),(-1,-1), 12),
        ]))
        story.append(score_tbl)
        story.append(Spacer(1, 6))

        # Risk factors
        risk_factors = ai_risk.get('risk_factors', [])
        if risk_factors:
            story.append(Paragraph('RISK FACTORS',
                style('RFTitle', fontSize=8, textColor=GREY, fontName='Helvetica-Bold', spaceBefore=8, spaceAfter=4)))
            impact_colors = {'high': RED_TEXT, 'medium': colors.HexColor('#b8960a'), 'low': colors.HexColor('#1a7a3c')}
            for f in risk_factors:
                ic = impact_colors.get(f.get('impact', 'medium'), GREY)
                rf_tbl = Table([[
                    Paragraph(f.get('factor', ''), style('RFactor', fontSize=10, textColor=DARK, fontName='Helvetica-Bold')),
                    Paragraph(f.get('impact', '').upper(), style('RImpact', fontSize=8, textColor=ic, fontName='Helvetica-Bold', alignment=TA_RIGHT)),
                ], [
                    Paragraph(f.get('detail', ''), style('RDetail', fontSize=9.5, textColor=GREY, fontName='Helvetica', colSpan=2)),
                    '',
                ]], colWidths=[page_w * 0.75, page_w * 0.25 - 0.3*cm])
                rf_tbl.setStyle(TableStyle([
                    ('BACKGROUND',   (0,0),(-1,-1), LIGHT_GREY),
                    ('BOX',          (0,0),(-1,-1), 0.5, BORDER),
                    ('SPAN',         (0,1),(1,1)),
                    ('TOPPADDING',   (0,0),(-1,-1), 5),
                    ('BOTTOMPADDING',(0,0),(-1,-1), 5),
                    ('LEFTPADDING',  (0,0),(-1,-1), 10),
                    ('RIGHTPADDING', (0,0),(-1,-1), 10),
                ]))
                story.append(rf_tbl)
                story.append(Spacer(1, 4))

        # Conditions to monitor
        conds_to_monitor = ai_risk.get('conditions_to_monitor', [])
        if conds_to_monitor:
            story.append(Paragraph('CONDITIONS TO MONITOR',
                style('CTMTitle', fontSize=8, textColor=GREY, fontName='Helvetica-Bold', spaceBefore=8, spaceAfter=4)))
            cond_cells = [Table([[Paragraph(c, style('CTM', fontSize=9.5, textColor=RED_TEXT, fontName='Helvetica-Bold'))]],
                               colWidths=[len(c)*6 + 24]) for c in conds_to_monitor]
            if cond_cells:
                ctm_row = Table([cond_cells], colWidths=[len(c)*6 + 28 for c in conds_to_monitor], hAlign='LEFT')
                ctm_row.setStyle(TableStyle([
                    ('BACKGROUND',   (0,0),(-1,-1), RED_BG),
                    ('BOX',          (0,0),(-1,-1), 0.5, colors.HexColor('#f5c6c6')),
                    ('LEFTPADDING',  (0,0),(-1,-1), 8),
                    ('RIGHTPADDING', (0,0),(-1,-1), 8),
                    ('TOPPADDING',   (0,0),(-1,-1), 4),
                    ('BOTTOMPADDING',(0,0),(-1,-1), 4),
                ]))
                story.append(ctm_row)
                story.append(Spacer(1, 4))

        # Preventive actions
        prev_actions = ai_risk.get('preventive_actions', [])
        if prev_actions:
            story.append(Paragraph('PREVENTIVE ACTIONS',
                style('PATitle', fontSize=8, textColor=GREY, fontName='Helvetica-Bold', spaceBefore=8, spaceAfter=4)))
            for i, action in enumerate(prev_actions, 1):
                pa_tbl = Table([[
                    Paragraph(str(i), style('PANum', fontSize=9, textColor=GREEN, fontName='Helvetica-Bold', alignment=TA_CENTER)),
                    Paragraph(action, style('PAText', fontSize=10, textColor=DARK, fontName='Helvetica')),
                ]], colWidths=[0.6*cm, page_w - 0.9*cm])
                pa_tbl.setStyle(TableStyle([
                    ('BACKGROUND',   (0,0),(0,0), GREEN_LIGHT),
                    ('VALIGN',       (0,0),(-1,-1), 'MIDDLE'),
                    ('TOPPADDING',   (0,0),(-1,-1), 5),
                    ('BOTTOMPADDING',(0,0),(-1,-1), 5),
                    ('LEFTPADDING',  (0,0),(-1,-1), 8),
                    ('RIGHTPADDING', (0,0),(-1,-1), 8),
                ]))
                story.append(pa_tbl)
                story.append(Spacer(1, 3))

    story.append(Spacer(1, 16))
    story.append(HRFlowable(width='100%', thickness=0.5, color=BORDER, spaceAfter=8))
    now_str = datetime.now().strftime('%B %d, %Y at %I:%M %p')
    story.append(Paragraph(f'Generated on {now_str}  -  MMSU Medical Health Records  -  Confidential', footer_style))

    doc.build(story)
    buf.seek(0)
    safe_name = p['name'].replace(' ', '_')
    return Response(
        buf.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{safe_name}_medical_record.pdf"'}
    )

# ── GROQ AI PROXY ─────────────────────────────────────────────────────────────

_groq_client = None
_groq_lock   = Lock()

def get_groq_client():
    global _groq_client
    if _groq_client is None:
        with _groq_lock:
            if _groq_client is None:
                try:
                    from groq import Groq
                except ImportError:
                    return None
                api_key = os.environ.get('GROQ_API_KEY', '').strip()
                if not api_key:
                    return None
                _groq_client = Groq(api_key=api_key)
    return _groq_client


@app.route('/ai/models')
@login_required
def ai_models():
    """Return the configured Groq model for debugging."""
    api_key = os.environ.get('GROQ_API_KEY', '').strip()
    if not api_key:
        return jsonify({'error': 'GROQ_API_KEY not set'}), 500
    return jsonify({'model': 'llama-3.1-8b-instant', 'provider': 'Groq', 'status': 'configured'})


@app.route('/ai/suggest', methods=['POST'])
@login_required
@csrf_required
def ai_suggest():
    """Proxy Groq API calls so the API key is never exposed to the browser."""
    client = get_groq_client()
    if client is None:
        try:
            import groq  # noqa: F401
        except ImportError:
            return jsonify({'error': 'groq package not installed. Add groq to requirements.txt and redeploy.'}), 500
        return jsonify({'error': 'GROQ_API_KEY is not configured on the server. Add it in your Railway environment variables.'}), 500

    payload  = request.json or {}
    messages = payload.get('messages', [])
    if not messages:
        return jsonify({'error': 'No messages provided.'}), 400

    prompt_text = messages[0].get('content', '')
    if not prompt_text or not prompt_text.strip():
        return jsonify({'error': 'Prompt content is empty.'}), 400

    # Hard cap: prevent runaway token costs and prompt-injection via huge payloads.
    # 4 000 chars ≈ ~1 000 tokens, well within the model's context but enough for
    # any legitimate clinical suggestion request from this app.
    MAX_PROMPT_CHARS = 4_000
    if len(prompt_text) > MAX_PROMPT_CHARS:
        return jsonify({'error': f'Prompt exceeds maximum length of {MAX_PROMPT_CHARS} characters.'}), 400

    try:
        chat_completion = client.chat.completions.create(
            model='llama-3.1-8b-instant',
            messages=[
                {'role': 'system', 'content': 'You are a clinical assistant. You MUST respond with valid JSON only. No explanation, no markdown, no extra text -- just the raw JSON object.'},
                {'role': 'user',   'content': prompt_text},
            ],
            max_tokens=1024,
            temperature=0.3,
            response_format={'type': 'json_object'},
        )
        text = chat_completion.choices[0].message.content
        app.logger.debug('[ai_suggest] RAW MODEL OUTPUT: %s', text[:500])

        if not text:
            return jsonify({'error': 'Empty response from Groq model.'}), 502

        return jsonify({'content': [{'text': text}]})

    except Exception as e:
        app.logger.error('[ai_suggest] Groq API error: %s', e, exc_info=True)
        return jsonify({'error': str(e)}), 500

# ── PDF / EXCEL REPORTS ───────────────────────────────────────────────────────

def _report_colors():
    """Shared color palette for all reports."""
    from reportlab.lib import colors
    return {
        'GREEN':      colors.HexColor('#1a7a3c'),
        'GREEN_LIGHT':colors.HexColor('#e8f5ee'),
        'GOLD':       colors.HexColor('#d4b84a'),
        'DARK':       colors.HexColor('#111111'),
        'GREY':       colors.HexColor('#555555'),
        'LIGHT_GREY': colors.HexColor('#f6f7f6'),
        'BORDER':     colors.HexColor('#dde0dd'),
        'RED_BG':     colors.HexColor('#fdecea'),
        'RED_TEXT':   colors.HexColor('#c0392b'),
        'WHITE':      colors.white,
    }

def _report_header_footer(canvas_obj, doc, title, subtitle=''):
    """Draws page header + footer on every page."""
    if not REPORTLAB_AVAILABLE:
        return
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    C = _report_colors()
    w, h = doc.pagesize
    # Header bar
    canvas_obj.setFillColor(C['GREEN'])
    canvas_obj.rect(0, h - 2.2*cm, w, 2.2*cm, fill=1, stroke=0)
    canvas_obj.setFillColor(C['GOLD'])
    canvas_obj.setFont('Helvetica-Bold', 13)
    canvas_obj.drawString(1.8*cm, h - 1.3*cm, 'MMSU Medical · Health Records System')
    canvas_obj.setFillColor(colors.white)
    canvas_obj.setFont('Helvetica', 9)
    canvas_obj.drawRightString(w - 1.8*cm, h - 1.3*cm, title)
    if subtitle:
        canvas_obj.setFont('Helvetica', 8)
        canvas_obj.drawRightString(w - 1.8*cm, h - 1.85*cm, subtitle)
    # Footer
    canvas_obj.setFillColor(C['BORDER'])
    canvas_obj.rect(0, 0, w, 1.1*cm, fill=1, stroke=0)
    canvas_obj.setFillColor(C['GREY'])
    canvas_obj.setFont('Helvetica', 7.5)
    now_str = datetime.now().strftime('%B %d, %Y at %I:%M %p')
    canvas_obj.drawString(1.8*cm, 0.38*cm, f'Generated: {now_str}  ·  Confidential — For Clinic Use Only')
    canvas_obj.drawRightString(w - 1.8*cm, 0.38*cm, f'Page {canvas_obj.getPageNumber()}')


def _style(styles, name, **kw):
    from reportlab.lib.styles import ParagraphStyle
    return ParagraphStyle(name, parent=styles['Normal'], **kw)


def _section_title(text, styles):
    from reportlab.platypus import Paragraph
    from reportlab.lib import colors
    C = _report_colors()
    s = _style(styles, f'ST_{text[:8]}',
               fontSize=8, fontName='Helvetica-Bold',
               textColor=C['GREY'], spaceBefore=14, spaceAfter=5,
               textTransform='uppercase', letterSpacing=1.2)
    return Paragraph(text, s)


def _make_table(data, col_widths, stripe=True):
    """Generic styled table."""
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    C = _report_colors()
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    cmds = [
        ('BACKGROUND',    (0,0), (-1,0), C['GREEN']),
        ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,-1), 8.5),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [C['WHITE'], C['LIGHT_GREY']] if stripe else [C['WHITE']]),
        ('GRID',          (0,0), (-1,-1), 0.35, C['BORDER']),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING',   (0,0), (-1,-1), 8),
        ('RIGHTPADDING',  (0,0), (-1,-1), 8),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
    ]
    tbl.setStyle(TableStyle(cmds))
    return tbl


# ── 1. MONTHLY MEDICAL REPORT ─────────────────────────────────────────────────

@app.route('/reports/monthly')
@login_required
def report_monthly():
    if not REPORTLAB_AVAILABLE:
        return jsonify({'error': 'reportlab not installed'}), 500

    year  = request.args.get('year',  datetime.now().year,  type=int)
    month = request.args.get('month', datetime.now().month, type=int)

    with get_db() as conn:
        c = conn.cursor()
        # Visits this month
        c.execute('''
            SELECT v.visit_date, p.name, p.department, v.reason, v.notes
            FROM visits v JOIN personnel p ON p.id = v.personnel_id
            WHERE EXTRACT(YEAR FROM v.visit_date)=%s AND EXTRACT(MONTH FROM v.visit_date)=%s
            ORDER BY v.visit_date, p.name
        ''', (year, month))
        visits = c.fetchall()

        # All personnel stats
        c.execute('SELECT id, name, age, gender, blood, department, conditions FROM personnel')
        personnel = [{'id':r[0],'name':r[1],'age':r[2],'gender':r[3],'blood':r[4],'dept':r[5],'conds':r[6].split('|') if r[6] else []} for r in c.fetchall()]

        # Top reasons this month
        c.execute('''
            SELECT COALESCE(NULLIF(reason,''),'General Visit') AS r, COUNT(*) n
            FROM visits
            WHERE EXTRACT(YEAR FROM visit_date)=%s AND EXTRACT(MONTH FROM visit_date)=%s
            GROUP BY r ORDER BY n DESC LIMIT 8
        ''', (year, month))
        top_reasons = c.fetchall()

        # Department visit counts
        c.execute('''
            SELECT p.department, COUNT(*) n
            FROM visits v JOIN personnel p ON p.id=v.personnel_id
            WHERE EXTRACT(YEAR FROM v.visit_date)=%s AND EXTRACT(MONTH FROM v.visit_date)=%s
            GROUP BY p.department ORDER BY n DESC
        ''', (year, month))
        dept_visits = c.fetchall()

    month_name = datetime(year, month, 1).strftime('%B %Y')
    total_p    = len(personnel)
    high_risk  = sum(1 for p in personnel if any(c in HIGH_RISK_CONDITIONS for c in p['conds']))

    import calendar
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    C  = _report_colors()
    styles = getSampleStyleSheet()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.8*cm, rightMargin=1.8*cm,
                            topMargin=2.8*cm,  bottomMargin=1.8*cm)

    def header_footer(canvas_obj, d):
        _report_header_footer(canvas_obj, d, 'Monthly Medical Report', month_name)

    story = []
    # Title block
    story.append(Paragraph(f'Monthly Medical Report', _style(styles,'RT', fontSize=20, fontName='Helvetica-Bold', textColor=C['DARK'], spaceAfter=2)))
    story.append(Paragraph(month_name, _style(styles,'RS', fontSize=13, textColor=C['GREY'], spaceAfter=16)))
    story.append(HRFlowable(width='100%', thickness=1.5, color=C['GREEN'], spaceAfter=14))

    # KPI row
    kpi_data = [['Total Personnel','High Risk','Visits This Month','Departments'],
                [str(total_p), str(high_risk), str(len(visits)), str(len(dept_visits))]]
    kpi_tbl = _make_table(kpi_data, [None]*4)
    story.append(kpi_tbl)
    story.append(Spacer(1, 14))

    # Top visit reasons
    if top_reasons:
        story.append(_section_title('Top Visit Reasons This Month', styles))
        rdata = [['Reason','Visit Count']] + [[r, str(n)] for r,n in top_reasons]
        pw = A4[0] - 3.6*cm
        story.append(_make_table(rdata, [pw*0.75, pw*0.25]))
        story.append(Spacer(1, 12))

    # Dept visit breakdown
    if dept_visits:
        story.append(_section_title('Visits by Department', styles))
        ddata = [['Department','Visits']] + [[d or '—', str(n)] for d,n in dept_visits]
        pw = A4[0] - 3.6*cm
        story.append(_make_table(ddata, [pw*0.75, pw*0.25]))
        story.append(Spacer(1, 12))

    # Visit log
    story.append(_section_title(f'Full Visit Log — {month_name}', styles))
    if visits:
        pw = A4[0] - 3.6*cm
        vdata = [['Date','Personnel','Department','Reason','Notes']]
        for vdate, name, dept, reason, notes in visits:
            vdata.append([str(vdate), name, dept or '—', reason or 'General Visit', (notes or '')[:60]])
        story.append(_make_table(vdata, [pw*0.11, pw*0.2, pw*0.18, pw*0.27, pw*0.24]))
    else:
        story.append(Paragraph('No visits recorded for this month.', _style(styles,'NV', fontSize=11, textColor=C['GREY'])))

    doc.build(story, onFirstPage=header_footer, onLaterPages=header_footer)
    buf.seek(0)
    fname = f'monthly_report_{year}_{month:02d}.pdf'
    audit('REPORT_MONTHLY', f'{month_name}')
    return Response(buf.getvalue(), mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename="{fname}"'})


# ── 2. YEARLY SUMMARY REPORT ──────────────────────────────────────────────────

@app.route('/reports/yearly')
@login_required
def report_yearly():
    if not REPORTLAB_AVAILABLE:
        return jsonify({'error': 'reportlab not installed'}), 500

    year = request.args.get('year', datetime.now().year, type=int)

    with get_db() as conn:
        c = conn.cursor()
        c.execute('SELECT name, age, gender, blood, department, conditions FROM personnel')
        personnel = [{'name':r[0],'age':r[1],'gender':r[2],'blood':r[3],'dept':r[4],
                      'conds':r[5].split('|') if r[5] else []} for r in c.fetchall()]

        # Visits per month
        c.execute('''
            SELECT EXTRACT(MONTH FROM visit_date)::int AS m, COUNT(*) n
            FROM visits WHERE EXTRACT(YEAR FROM visit_date)=%s
            GROUP BY m ORDER BY m
        ''', (year,))
        visits_by_month = {r[0]: r[1] for r in c.fetchall()}

        # Total visits
        c.execute('SELECT COUNT(*) FROM visits WHERE EXTRACT(YEAR FROM visit_date)=%s', (year,))
        total_visits = c.fetchone()[0]

        # Top conditions
        c.execute('''
            SELECT COALESCE(NULLIF(reason,''),'General Visit'), COUNT(*)
            FROM visits WHERE EXTRACT(YEAR FROM visit_date)=%s
            GROUP BY 1 ORDER BY 2 DESC LIMIT 10
        ''', (year,))
        top_reasons = c.fetchall()

        # Dept breakdown
        c.execute('''
            SELECT p.department, COUNT(DISTINCT p.id) personnel,
                   SUM(CASE WHEN v.id IS NOT NULL THEN 1 ELSE 0 END) visits
            FROM personnel p LEFT JOIN visits v ON v.personnel_id=p.id
              AND EXTRACT(YEAR FROM v.visit_date)=%s
            GROUP BY p.department ORDER BY personnel DESC
        ''', (year,))
        dept_rows = c.fetchall()

    import calendar
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.lib.pagesizes import A4
    C  = _report_colors()
    styles = getSampleStyleSheet()

    total_p   = len(personnel)
    high_risk = sum(1 for p in personnel if any(c in HIGH_RISK_CONDITIONS for c in p['conds']))
    male      = sum(1 for p in personnel if p['gender'] == 'Male')
    female    = sum(1 for p in personnel if p['gender'] == 'Female')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.8*cm, rightMargin=1.8*cm,
                            topMargin=2.8*cm, bottomMargin=1.8*cm)

    def hf(cv, d): _report_header_footer(cv, d, 'Yearly Summary Report', str(year))

    story = []
    story.append(Paragraph('Yearly Summary Report', _style(styles,'RT', fontSize=20, fontName='Helvetica-Bold', textColor=C['DARK'], spaceAfter=2)))
    story.append(Paragraph(str(year), _style(styles,'RS', fontSize=13, textColor=C['GREY'], spaceAfter=16)))
    story.append(HRFlowable(width='100%', thickness=1.5, color=C['GREEN'], spaceAfter=14))

    pw = A4[0] - 3.6*cm
    kpi = [['Total Personnel','High Risk','Male','Female','Total Visits'],
           [str(total_p), str(high_risk), str(male), str(female), str(total_visits)]]
    story.append(_make_table(kpi, [pw/5]*5))
    story.append(Spacer(1, 14))

    # Monthly visits table
    story.append(_section_title('Monthly Visit Volume', styles))
    MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    mdata = [['Month','Visits']] + [[MONTHS[i], str(visits_by_month.get(i+1, 0))] for i in range(12)]
    story.append(_make_table(mdata, [pw*0.6, pw*0.4]))
    story.append(Spacer(1, 12))

    # Department table
    story.append(_section_title('Department Overview', styles))
    ddata = [['Department','Personnel','Visits (Year)','High Risk']]
    for dept, pop, vis in dept_rows:
        dept_hr = sum(1 for p in personnel if p['dept'] == dept and any(c in HIGH_RISK_CONDITIONS for c in p['conds']))
        ddata.append([dept or 'Unassigned', str(pop), str(vis or 0), str(dept_hr)])
    story.append(_make_table(ddata, [pw*0.4, pw*0.2, pw*0.2, pw*0.2]))
    story.append(Spacer(1, 12))

    # Top reasons
    if top_reasons:
        story.append(_section_title('Top 10 Visit Reasons', styles))
        rdata = [['Reason','Count']] + [[r, str(n)] for r,n in top_reasons]
        story.append(_make_table(rdata, [pw*0.75, pw*0.25]))

    doc.build(story, onFirstPage=hf, onLaterPages=hf)
    buf.seek(0)
    audit('REPORT_YEARLY', str(year))
    return Response(buf.getvalue(), mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename="yearly_report_{year}.pdf"'})


# ── 3. DEPARTMENT REPORT ──────────────────────────────────────────────────────

@app.route('/reports/department')
@login_required
def report_department():
    if not REPORTLAB_AVAILABLE:
        return jsonify({'error': 'reportlab not installed'}), 500

    dept = request.args.get('dept', '')

    with get_db() as conn:
        c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        if dept:
            c.execute('SELECT id,name,age,gender,blood,department,phone,address,conditions FROM personnel WHERE department=%s ORDER BY name', (dept,))
        else:
            c.execute('SELECT id,name,age,gender,blood,department,phone,address,conditions FROM personnel ORDER BY department,name')
        rows = c.fetchall()
        personnel = [row_to_person(r) for r in rows]

        c2 = conn.cursor()
        c2.execute('''
            SELECT p.department, COUNT(DISTINCT v.id) visits
            FROM personnel p LEFT JOIN visits v ON v.personnel_id=p.id
            WHERE (%s='' OR p.department=%s)
            GROUP BY p.department ORDER BY visits DESC
        ''', (dept, dept))
        dept_visit_counts = {r[0]: r[1] for r in c2.fetchall()}

    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.lib.pagesizes import A4
    C  = _report_colors()
    styles = getSampleStyleSheet()

    title_str = dept if dept else 'All Departments'
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.8*cm, rightMargin=1.8*cm,
                            topMargin=2.8*cm, bottomMargin=1.8*cm)

    def hf(cv, d): _report_header_footer(cv, d, 'Department Report', title_str)

    story = []
    story.append(Paragraph('Department Report', _style(styles,'RT', fontSize=20, fontName='Helvetica-Bold', textColor=C['DARK'], spaceAfter=2)))
    story.append(Paragraph(title_str, _style(styles,'RS', fontSize=13, textColor=C['GREY'], spaceAfter=16)))
    story.append(HRFlowable(width='100%', thickness=1.5, color=C['GREEN'], spaceAfter=14))

    pw = A4[0] - 3.6*cm

    # Group by department
    from collections import defaultdict
    by_dept = defaultdict(list)
    for p in personnel:
        by_dept[p['department'] or 'Unassigned'].append(p)

    for dname, members in sorted(by_dept.items()):
        story.append(_section_title(dname, styles))
        high = sum(1 for p in members if any(c in HIGH_RISK_CONDITIONS for c in p['conditions']))
        info = [['Personnel','High Risk','Total Visits'],
                [str(len(members)), str(high), str(dept_visit_counts.get(dname, 0))]]
        story.append(_make_table(info, [pw/3]*3))
        story.append(Spacer(1, 6))

        pdata = [['Name','Age','Gender','Blood','Risk','Conditions']]
        for p in members:
            risk = '⚠ High' if any(c in HIGH_RISK_CONDITIONS for c in p['conditions']) else '✓ Normal'
            conds = ', '.join(p['conditions'][:3]) + ('…' if len(p['conditions'])>3 else '') if p['conditions'] else '—'
            pdata.append([p['name'], str(p['age'] or '—'), p['gender'], p['blood'] or '—', risk, conds])
        story.append(_make_table(pdata, [pw*0.22, pw*0.07, pw*0.08, pw*0.08, pw*0.12, pw*0.43]))
        story.append(Spacer(1, 14))

    doc.build(story, onFirstPage=hf, onLaterPages=hf)
    buf.seek(0)
    safe_dept = dept.replace(' ','_') if dept else 'all'
    audit('REPORT_DEPT', title_str)
    return Response(buf.getvalue(), mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename="dept_report_{safe_dept}.pdf"'})


# ── 4. CONSULTATION HISTORY (PDF) ─────────────────────────────────────────────

@app.route('/reports/consultation')
@login_required
def report_consultation():
    if not REPORTLAB_AVAILABLE:
        return jsonify({'error': 'reportlab not installed'}), 500

    pid  = request.args.get('pid',  type=int)
    dept = request.args.get('dept', '')
    from_date = request.args.get('from', '')
    to_date   = request.args.get('to', '')

    with get_db() as conn:
        c = conn.cursor()
        query = '''
            SELECT v.id, p.name, p.department, p.gender, p.age,
                   v.visit_date, v.reason, v.notes
            FROM visits v JOIN personnel p ON p.id = v.personnel_id
            WHERE 1=1
        '''
        params = []
        if pid:
            query += ' AND p.id=%s'; params.append(pid)
        if dept:
            query += ' AND p.department=%s'; params.append(dept)
        if from_date:
            query += ' AND v.visit_date>=%s'; params.append(from_date)
        if to_date:
            query += ' AND v.visit_date<=%s'; params.append(to_date)
        query += ' ORDER BY v.visit_date DESC, p.name'
        c.execute(query, params)
        visits = c.fetchall()

    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.lib.pagesizes import A4
    C  = _report_colors()
    styles = getSampleStyleSheet()

    subtitle_parts = []
    if dept: subtitle_parts.append(dept)
    if from_date: subtitle_parts.append(f'From {from_date}')
    if to_date:   subtitle_parts.append(f'To {to_date}')
    subtitle = ' · '.join(subtitle_parts) if subtitle_parts else 'All Records'

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.8*cm, rightMargin=1.8*cm,
                            topMargin=2.8*cm, bottomMargin=1.8*cm)

    def hf(cv, d): _report_header_footer(cv, d, 'Consultation History', subtitle)

    story = []
    story.append(Paragraph('Consultation History', _style(styles,'RT', fontSize=20, fontName='Helvetica-Bold', textColor=C['DARK'], spaceAfter=2)))
    story.append(Paragraph(subtitle, _style(styles,'RS', fontSize=12, textColor=C['GREY'], spaceAfter=16)))
    story.append(HRFlowable(width='100%', thickness=1.5, color=C['GREEN'], spaceAfter=14))

    pw = A4[0] - 3.6*cm
    story.append(Paragraph(f'Total consultations: {len(visits)}', _style(styles,'TC', fontSize=11, textColor=C['GREY'], spaceAfter=10)))

    if visits:
        vdata = [['Date','Personnel','Dept','Age','Gender','Reason','Notes']]
        for _, name, dept_v, gender, age, vdate, reason, notes in visits:
            vdata.append([str(vdate), name, dept_v or '—', str(age or '—'), gender or '—',
                          (reason or 'General Visit')[:35], (notes or '')[:50]])
        story.append(_make_table(vdata, [pw*0.1, pw*0.17, pw*0.13, pw*0.05, pw*0.07, pw*0.24, pw*0.24]))
    else:
        story.append(Paragraph('No consultation records found for the selected filters.', _style(styles,'NR', fontSize=11, textColor=C['GREY'])))

    doc.build(story, onFirstPage=hf, onLaterPages=hf)
    buf.seek(0)
    audit('REPORT_CONSULTATION', subtitle)
    return Response(buf.getvalue(), mimetype='application/pdf',
                    headers={'Content-Disposition': 'attachment; filename="consultation_history.pdf"'})


# ── 5. MEDICINE INVENTORY REPORT (EXCEL) ─────────────────────────────────────

@app.route('/reports/medicine-inventory')
@login_required
def report_medicine_inventory():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'openpyxl not installed. Run: pip install openpyxl'}), 500

    with get_db() as conn:
        c = conn.cursor()
        c.execute('SELECT name, age, gender, blood, department, conditions FROM personnel')
        personnel = [{'name':r[0],'age':r[1],'gender':r[2],'blood':r[3],'dept':r[4],
                      'conds':r[5].split('|') if r[5] else []} for r in c.fetchall()]

    # Count condition occurrences
    cond_counts = defaultdict(int)
    cond_dept   = defaultdict(lambda: defaultdict(int))
    for p in personnel:
        for cond in p['conds']:
            if cond:
                cond_counts[cond] += 1
                cond_dept[cond][p['dept'] or 'Unassigned'] += 1

    total_p = len(personnel)

    # Medicine suggestions per condition (static, curated)
    MED_DB = {
        'Hypertension':   ['Amlodipine 5mg','Losartan 50mg','Hydrochlorothiazide 25mg'],
        'Diabetes':       ['Metformin 500mg','Glibenclamide 5mg','Insulin (Regular)'],
        'Asthma':         ['Salbutamol Inhaler','Ipratropium Inhaler','Prednisolone 5mg'],
        'Heart Disease':  ['Aspirin 81mg','Atorvastatin 40mg','Metoprolol 50mg'],
        'Tuberculosis':   ['Isoniazid 300mg','Rifampicin 600mg','Pyrazinamide 500mg','Ethambutol 400mg'],
        'Cancer':         ['(Refer to Oncology)','Ondansetron 4mg','Dexamethasone 4mg'],
        'HIV/AIDS':       ['(Refer to Infectious Disease)','Cotrimoxazole 960mg'],
        'Epilepsy':       ['Phenobarbital 30mg','Carbamazepine 200mg','Valproic Acid 500mg'],
    }

    wb = openpyxl.Workbook()

    # ── colours / helpers ──
    HDR_FILL  = PatternFill('solid', start_color='1a7a3c')
    GOLD_FILL = PatternFill('solid', start_color='d4b84a')
    RISK_FILL = PatternFill('solid', start_color='fdecea')
    ALT_FILL  = PatternFill('solid', start_color='f6f7f6')
    WHITE     = PatternFill('solid', start_color='FFFFFF')
    thin      = Side(style='thin', color='DDDDDD')
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(ws, row, col, text, bold=True, color='FFFFFF'):
        c = ws.cell(row=row, column=col, value=text)
        c.font      = Font(bold=bold, color=color, name='Arial', size=9)
        c.fill      = HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = bdr
        return c

    def data_cell(ws, row, col, value, bold=False, fill=None, align='left', color='111111'):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(bold=bold, name='Arial', size=9, color=color)
        c.fill      = fill or WHITE
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        c.border    = bdr
        return c

    # ── SHEET 1: Condition Inventory ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Condition Inventory'
    ws1.row_dimensions[1].height = 30
    ws1.freeze_panes = 'A2'

    headers1 = ['Condition','Affected','% of Personnel','Risk Level','Suggested Medicines']
    col_w1   = [28, 12, 16, 14, 55]
    for i, (h, w) in enumerate(zip(headers1, col_w1), 1):
        hdr_cell(ws1, 1, i, h)
        ws1.column_dimensions[get_column_letter(i)].width = w

    sorted_conds = sorted(cond_counts.items(), key=lambda x: -x[1])
    for r, (cond, cnt) in enumerate(sorted_conds, 2):
        is_hr  = cond in HIGH_RISK_CONDITIONS
        fill   = RISK_FILL if is_hr else (ALT_FILL if r%2==0 else WHITE)
        pct    = cnt / total_p * 100 if total_p else 0
        meds   = ', '.join(MED_DB.get(cond, ['See clinical guidelines']))
        data_cell(ws1, r, 1, cond, bold=is_hr, fill=fill, color='c0392b' if is_hr else '111111')
        data_cell(ws1, r, 2, cnt, align='center', fill=fill)
        data_cell(ws1, r, 3, f'{pct:.1f}%', align='center', fill=fill)
        data_cell(ws1, r, 4, '⚠ High Risk' if is_hr else 'Standard', align='center', fill=fill,
                  color='c0392b' if is_hr else '1a7a3c', bold=is_hr)
        data_cell(ws1, r, 5, meds, fill=fill)

    ws1.auto_filter.ref = f'A1:E{len(sorted_conds)+1}'

    # ── SHEET 2: Department Breakdown ─────────────────────────────────────────
    ws2 = wb.create_sheet('By Department')
    ws2.freeze_panes = 'A2'
    ws2.row_dimensions[1].height = 30

    # Gather all conditions per dept
    dept_cond: dict = defaultdict(lambda: defaultdict(int))
    dept_total: dict = defaultdict(int)
    for p in personnel:
        d = p['dept'] or 'Unassigned'
        dept_total[d] += 1
        for cond in p['conds']:
            if cond:
                dept_cond[d][cond] += 1

    # Unique conditions seen
    all_conds_sorted = sorted(cond_counts.keys())
    h2 = ['Department','Total Personnel'] + all_conds_sorted
    cw2 = [24, 16] + [max(len(c)//1.5, 10) for c in all_conds_sorted]
    for i, (h, w) in enumerate(zip(h2, cw2), 1):
        hdr_cell(ws2, 1, i, h)
        ws2.column_dimensions[get_column_letter(i)].width = w

    for r, (dept_name, tot) in enumerate(sorted(dept_total.items()), 2):
        fill = ALT_FILL if r%2==0 else WHITE
        data_cell(ws2, r, 1, dept_name, bold=True, fill=fill)
        data_cell(ws2, r, 2, tot, align='center', fill=fill)
        for ci, cond in enumerate(all_conds_sorted, 3):
            val = dept_cond[dept_name].get(cond, 0)
            is_hr = cond in HIGH_RISK_CONDITIONS
            cf = RISK_FILL if (val > 0 and is_hr) else fill
            data_cell(ws2, r, ci, val if val else '', align='center', fill=cf,
                      color='c0392b' if (val > 0 and is_hr) else '555555')

    ws2.auto_filter.ref = f'A1:{get_column_letter(len(h2))}1'

    # ── SHEET 3: Medicine Needs Estimate ──────────────────────────────────────
    ws3 = wb.create_sheet('Medicine Needs')
    ws3.freeze_panes = 'A2'
    ws3.row_dimensions[1].height = 30

    h3 = ['Medicine','For Condition','Risk Level','Patients Affected','Est. Monthly Units','Notes']
    cw3 = [32, 20, 14, 16, 18, 40]
    for i, (h, w) in enumerate(zip(h3, cw3), 1):
        hdr_cell(ws3, 1, i, h)
        ws3.column_dimensions[get_column_letter(i)].width = w

    r = 2
    for cond, meds in MED_DB.items():
        affected = cond_counts.get(cond, 0)
        if affected == 0:
            continue
        is_hr = cond in HIGH_RISK_CONDITIONS
        for med in meds:
            fill = RISK_FILL if is_hr else (ALT_FILL if r%2==0 else WHITE)
            monthly = affected * 30  # rough 1 unit/day estimate
            data_cell(ws3, r, 1, med, bold=True, fill=fill)
            data_cell(ws3, r, 2, cond, fill=fill)
            data_cell(ws3, r, 3, '⚠ High Risk' if is_hr else 'Standard', align='center', fill=fill,
                      color='c0392b' if is_hr else '1a7a3c', bold=is_hr)
            data_cell(ws3, r, 4, affected, align='center', fill=fill)
            data_cell(ws3, r, 5, monthly, align='center', fill=fill)
            data_cell(ws3, r, 6, 'Estimate only — consult pharmacist for actual procurement.', fill=fill,
                      color='888888')
            r += 1

    # ── SHEET 4: Summary Dashboard ────────────────────────────────────────────
    ws4 = wb.create_sheet('Summary')
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions['A'].width = 28
    ws4.column_dimensions['B'].width = 18

    def title_cell(ws, row, col, text, size=14):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(bold=True, name='Arial', size=size, color='1a7a3c')
        c.alignment = Alignment(vertical='center')

    title_cell(ws4, 1, 1, 'MMSU Medical — Medicine Inventory Report', 14)
    ws4.cell(row=2, column=1, value=f'Generated: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}').font = Font(name='Arial', size=9, color='888888')

    summary_rows = [
        ('', ''),
        ('KPI', 'Value'),
        ('Total Personnel', total_p),
        ('Unique Conditions Tracked', len(cond_counts)),
        ('High-Risk Conditions', sum(1 for c in cond_counts if c in HIGH_RISK_CONDITIONS)),
        ('Personnel with ≥1 Condition', sum(1 for p in personnel if p['conds'])),
        ('Personnel — High Risk', sum(1 for p in personnel if any(c in HIGH_RISK_CONDITIONS for c in p['conds']))),
        ('Most Common Condition', sorted_conds[0][0] if sorted_conds else '—'),
        ('Least Common Condition', sorted_conds[-1][0] if sorted_conds else '—'),
    ]
    for sr, (label, val) in enumerate(summary_rows, 4):
        if label == 'KPI':
            hdr_cell(ws4, sr, 1, label); hdr_cell(ws4, sr, 2, 'Value')
        else:
            fill = ALT_FILL if sr%2==0 else WHITE
            data_cell(ws4, sr, 1, label, bold=True, fill=fill)
            data_cell(ws4, sr, 2, val, align='center', fill=fill)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    audit('REPORT_MEDICINE_INVENTORY', f'{len(cond_counts)} conditions')
    return Response(buf.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename="medicine_inventory_report.xlsx"'})

# ── TREND ANALYTICS ───────────────────────────────────────────────────────────

@app.route('/analytics/trends')
@login_required
def analytics_trends():
    """
    All trend data in one efficient query set.

    Query params:
      period  – '6m' | '12m' | 'all' (default '12m')
    """
    period = request.args.get('period', '12m')

    now = datetime.now()
    if period == '6m':
        cutoff = datetime(now.year, now.month, 1) - timedelta(days=5*31)
        cutoff = cutoff.replace(day=1)
    elif period == '12m':
        cutoff = datetime(now.year, now.month, 1) - timedelta(days=11*31)
        cutoff = cutoff.replace(day=1)
    else:
        cutoff = None

    with get_db() as conn:
        c = conn.cursor()

        # ── 1. Monthly visit volume ──────────────────────────────────────────
        if cutoff:
            c.execute("""
                SELECT TO_CHAR(visit_date, 'YYYY-MM') AS mo, COUNT(*) AS cnt
                FROM visits
                WHERE visit_date >= %s
                GROUP BY mo ORDER BY mo
            """, (cutoff.date(),))
        else:
            c.execute("""
                SELECT TO_CHAR(visit_date, 'YYYY-MM') AS mo, COUNT(*) AS cnt
                FROM visits
                GROUP BY mo ORDER BY mo
            """)
        monthly_visits = [{'month': r[0], 'count': r[1]} for r in c.fetchall()]

        # ── 2. Day-of-week distribution ──────────────────────────────────────
        if cutoff:
            c.execute("""
                SELECT EXTRACT(DOW FROM visit_date)::int AS dow, COUNT(*) AS cnt
                FROM visits WHERE visit_date >= %s
                GROUP BY dow ORDER BY dow
            """, (cutoff.date(),))
        else:
            c.execute("""
                SELECT EXTRACT(DOW FROM visit_date)::int AS dow, COUNT(*) AS cnt
                FROM visits GROUP BY dow ORDER BY dow
            """)
        dow_raw = {r[0]: r[1] for r in c.fetchall()}
        day_names = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
        dow_distribution = [{'day': day_names[i], 'count': dow_raw.get(i, 0)} for i in range(7)]

        # ── 3. Top visit reasons (monthly breakdown for top 5) ───────────────
        if cutoff:
            c.execute("""
                SELECT reason, COUNT(*) AS cnt FROM visits
                WHERE visit_date >= %s AND reason IS NOT NULL AND reason != ''
                GROUP BY reason ORDER BY cnt DESC LIMIT 7
            """, (cutoff.date(),))
        else:
            c.execute("""
                SELECT reason, COUNT(*) AS cnt FROM visits
                WHERE reason IS NOT NULL AND reason != ''
                GROUP BY reason ORDER BY cnt DESC LIMIT 7
            """)
        top_reasons = [{'reason': r[0], 'count': r[1]} for r in c.fetchall()]

        # ── 4. Visit volume by department (month × dept) ─────────────────────
        if cutoff:
            c.execute("""
                SELECT TO_CHAR(v.visit_date, 'YYYY-MM') AS mo,
                       COALESCE(p.department, 'Unknown') AS dept,
                       COUNT(*) AS cnt
                FROM visits v JOIN personnel p ON p.id = v.personnel_id
                WHERE v.visit_date >= %s
                GROUP BY mo, dept ORDER BY mo, cnt DESC
            """, (cutoff.date(),))
        else:
            c.execute("""
                SELECT TO_CHAR(v.visit_date, 'YYYY-MM') AS mo,
                       COALESCE(p.department, 'Unknown') AS dept,
                       COUNT(*) AS cnt
                FROM visits v JOIN personnel p ON p.id = v.personnel_id
                GROUP BY mo, dept ORDER BY mo, cnt DESC
            """)
        dept_monthly_raw = c.fetchall()

        # Top 5 departments by total visits
        dept_totals = {}
        for mo, dept, cnt in dept_monthly_raw:
            dept_totals[dept] = dept_totals.get(dept, 0) + cnt
        top_depts = [d for d, _ in sorted(dept_totals.items(), key=lambda x: -x[1])[:5]]

        # Build month × dept matrix
        dept_month_map = {}
        for mo, dept, cnt in dept_monthly_raw:
            if dept in top_depts:
                dept_month_map.setdefault(dept, {})[mo] = cnt

        dept_monthly = {dept: dept_month_map.get(dept, {}) for dept in top_depts}

        # ── 5. Condition prevalence snapshot (current, not time-series) ──────
        c.execute("""
            SELECT conditions FROM personnel WHERE conditions IS NOT NULL AND conditions != ''
        """)
        cond_counts = {}
        for (conds_str,) in c.fetchall():
            for cond in conds_str.split('|'):
                cond = cond.strip()
                if cond:
                    cond_counts[cond] = cond_counts.get(cond, 0) + 1
        top_conditions = sorted(cond_counts.items(), key=lambda x: -x[1])[:10]

        # ── 6. Summary stats ─────────────────────────────────────────────────
        c.execute('SELECT COUNT(*) FROM visits' + (' WHERE visit_date >= %s' if cutoff else ''),
                  (cutoff.date(),) if cutoff else ())
        total_visits = c.fetchone()[0]

        c.execute('SELECT COUNT(*) FROM personnel')
        total_personnel = c.fetchone()[0]

        # Returning visitors (visited more than once in period)
        if cutoff:
            c.execute("""
                SELECT COUNT(*) FROM (
                    SELECT personnel_id FROM visits WHERE visit_date >= %s
                    GROUP BY personnel_id HAVING COUNT(*) > 1
                ) t
            """, (cutoff.date(),))
        else:
            c.execute("""
                SELECT COUNT(*) FROM (
                    SELECT personnel_id FROM visits
                    GROUP BY personnel_id HAVING COUNT(*) > 1
                ) t
            """)
        returning = c.fetchone()[0]

        # Unique visitors in period
        if cutoff:
            c.execute("SELECT COUNT(DISTINCT personnel_id) FROM visits WHERE visit_date >= %s", (cutoff.date(),))
        else:
            c.execute("SELECT COUNT(DISTINCT personnel_id) FROM visits")
        unique_visitors = c.fetchone()[0]

    return jsonify({
        'period': period,
        'summary': {
            'total_visits': total_visits,
            'total_personnel': total_personnel,
            'unique_visitors': unique_visitors,
            'returning_visitors': returning,
        },
        'monthly_visits': monthly_visits,
        'dow_distribution': dow_distribution,
        'top_reasons': top_reasons,
        'dept_monthly': dept_monthly,
        'top_conditions': [{'condition': c, 'count': n} for c, n in top_conditions],
    })


# ── NOTIFICATIONS ─────────────────────────────────────────────────────────────

@app.route('/notifications')
@login_required
def get_notifications():
    with get_db() as conn:
        c = conn.cursor()
        c.execute(
            'SELECT id, level, title, body, read, created_at FROM notifications ORDER BY created_at DESC LIMIT 100'
        )
        rows = c.fetchall()
        c.execute('SELECT COUNT(*) FROM notifications WHERE read = FALSE')
        unread = c.fetchone()[0]
    return jsonify({
        'unread': unread,
        'data': [
            {'id': r[0], 'level': r[1], 'title': r[2], 'body': r[3], 'read': r[4], 'created_at': str(r[5])}
            for r in rows
        ],
    })


@app.route('/notifications/read-all', methods=['POST'])
@login_required
def mark_notifications_read():
    with get_db() as conn:
        conn.cursor().execute('UPDATE notifications SET read = TRUE WHERE read = FALSE')
    return jsonify({'ok': True})


@app.route('/notifications/delete-all', methods=['DELETE'])
@login_required
@csrf_required
def delete_notifications():
    with get_db() as conn:
        conn.cursor().execute('DELETE FROM notifications')
    return jsonify({'ok': True})


@app.route('/notifications/<int:nid>/read', methods=['POST'])
@login_required
def mark_notification_read(nid):
    with get_db() as conn:
        conn.cursor().execute('UPDATE notifications SET read = TRUE WHERE id = %s', (nid,))
    return jsonify({'ok': True})

# ── BACKUP / RESTORE ───────────────────────────────────────────────────────────

@app.route('/backup')
@login_required
def create_backup():
    """Download a full JSON backup of all data (personnel + visits + departments)."""
    with get_db() as conn:
        c = conn.cursor()
        c.execute('SELECT id, name, age, gender, blood, department, phone, address, conditions FROM personnel ORDER BY id')
        personnel = [
            {'id': r[0], 'name': r[1], 'age': r[2], 'gender': r[3], 'blood': r[4],
             'department': r[5], 'phone': r[6], 'address': r[7], 'conditions': r[8]}
            for r in c.fetchall()
        ]
        c.execute(
            'SELECT v.id, v.personnel_id, p.name, v.visit_date, v.reason, v.notes, v.created_at'
            ' FROM visits v JOIN personnel p ON p.id = v.personnel_id ORDER BY v.id'
        )
        visits = [
            {'id': r[0], 'personnel_id': r[1], 'personnel_name': r[2],
             'visit_date': str(r[3]), 'reason': r[4], 'notes': r[5], 'created_at': str(r[6])}
            for r in c.fetchall()
        ]
        c.execute('SELECT id, name FROM departments ORDER BY name')
        departments = [{'id': r[0], 'name': r[1]} for r in c.fetchall()]

    payload = {
        'backup_version': 1,
        'created_at': datetime.now().isoformat(),
        'counts': {
            'personnel': len(personnel),
            'visits': len(visits),
            'departments': len(departments),
        },
        'personnel': personnel,
        'visits': visits,
        'departments': departments,
    }

    audit('BACKUP', f'{len(personnel)} personnel, {len(visits)} visits, {len(departments)} departments')
    notify('Backup created', f'{len(personnel)} personnel, {len(visits)} visits, {len(departments)} departments', 'info')

    filename = f'mmsu_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
    return Response(
        json.dumps(payload, indent=2),
        mimetype='application/json',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@app.route('/restore', methods=['POST'])
@login_required
@csrf_required
def restore_backup():
    """Restore from a JSON backup. Personnel, visits, departments are fully replaced."""
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'No file provided'}), 400
    try:
        payload = json.loads(file.read().decode('utf-8'))
    except (UnicodeDecodeError, json.JSONDecodeError) as e:
        return jsonify({'error': f'Invalid JSON file: {e}'}), 400

    if payload.get('backup_version') != 1:
        return jsonify({'error': 'Unrecognised backup format or version'}), 400

    personnel   = payload.get('personnel', [])
    visits      = payload.get('visits', [])
    departments = payload.get('departments', [])

    # ── Validate all three entities before opening any transaction ────────────
    # Collecting every error up front means the user sees all problems at once,
    # and we never reach the DELETE statements with bad data.
    val_errors = []

    if not isinstance(personnel, list):
        val_errors.append('personnel must be a list')
    else:
        for i, p in enumerate(personnel):
            if not isinstance(p, dict):
                val_errors.append(f'personnel[{i}]: must be an object')
                continue
            if not isinstance(p.get('id'), int):
                val_errors.append(f'personnel[{i}]: missing or invalid "id" (integer required)')
            if not p.get('name') or not isinstance(p.get('name'), str):
                val_errors.append(f'personnel[{i}]: missing or invalid "name"')

    if not isinstance(visits, list):
        val_errors.append('visits must be a list')
    else:
        for i, v in enumerate(visits):
            if not isinstance(v, dict):
                val_errors.append(f'visits[{i}]: must be an object')
                continue
            if not isinstance(v.get('id'), int):
                val_errors.append(f'visits[{i}]: missing or invalid "id" (integer required)')
            if not isinstance(v.get('personnel_id'), int):
                val_errors.append(f'visits[{i}]: missing or invalid "personnel_id" (integer required)')
            if not v.get('visit_date'):
                val_errors.append(f'visits[{i}]: missing "visit_date"')

    if not isinstance(departments, list):
        val_errors.append('departments must be a list')
    else:
        for i, d in enumerate(departments):
            if not isinstance(d, dict):
                val_errors.append(f'departments[{i}]: must be an object')
                continue
            if not isinstance(d.get('id'), int):
                val_errors.append(f'departments[{i}]: missing or invalid "id" (integer required)')
            if not d.get('name') or not isinstance(d.get('name'), str):
                val_errors.append(f'departments[{i}]: missing or invalid "name"')

    if val_errors:
        return jsonify({
            'error': f'{len(val_errors)} validation error(s) in backup file.',
            'validation_errors': val_errors,
        }), 400

    with get_db() as conn:
        c = conn.cursor()
        c.execute('DELETE FROM departments')
        if departments:
            psycopg2.extras.execute_batch(
                c,
                'INSERT INTO departments (id, name) VALUES (%s, %s)',
                [(d['id'], d['name']) for d in departments],
            )
            c.execute("SELECT setval('departments_id_seq', (SELECT COALESCE(MAX(id),0) FROM departments))")
        c.execute('DELETE FROM personnel')
        if personnel:
            psycopg2.extras.execute_batch(
                c,
                'INSERT INTO personnel (id, name, age, gender, blood, department, phone, address, conditions)'
                ' VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                [(p['id'], p.get('name'), p.get('age'), p.get('gender'), p.get('blood'),
                  p.get('department'), p.get('phone'), p.get('address'), p.get('conditions'))
                 for p in personnel],
            )
            c.execute("SELECT setval('personnel_id_seq', (SELECT COALESCE(MAX(id),0) FROM personnel))")
        if visits:
            psycopg2.extras.execute_batch(
                c,
                'INSERT INTO visits (id, personnel_id, visit_date, reason, notes)'
                ' VALUES (%s,%s,%s,%s,%s)',
                [(v['id'], v['personnel_id'], v['visit_date'], v.get('reason'), v.get('notes'))
                 for v in visits],
            )
            c.execute("SELECT setval('visits_id_seq', (SELECT COALESCE(MAX(id),0) FROM visits))")

    audit('RESTORE', f'{len(personnel)} personnel, {len(visits)} visits, {len(departments)} departments')
    notify('Backup restored',
           f'{len(personnel)} personnel, {len(visits)} visits, {len(departments)} departments',
           'warning')
    return jsonify({
        'message': (f'Restore complete — {len(personnel)} personnel, '
                    f'{len(visits)} visits, {len(departments)} departments restored.'),
    })


# ── CHANGE PASSWORD ────────────────────────────────────────────────────────────────────────────────

@app.route('/settings/change-password', methods=['POST'])
@login_required
@csrf_required
def change_password():
    data = request.json or {}
    current = data.get('current_password', '')
    new_pw  = data.get('new_password', '')
    confirm = data.get('confirm_password', '')

    if not all([current, new_pw, confirm]):
        return jsonify({'error': 'All fields are required'}), 400
    if not check_password_hash(_get_password_hash(), current):
        return jsonify({'error': 'Current password is incorrect'}), 401
    if new_pw != confirm:
        return jsonify({'error': 'New passwords do not match'}), 400
    if len(new_pw) < 8:
        return jsonify({'error': 'Password must be at least 8 characters'}), 400

    new_hash = generate_password_hash(new_pw)
    with get_db() as conn:
        c = conn.cursor()
        c.execute("""
            INSERT INTO app_settings (key, value) VALUES ('admin_password_hash', %s)
            ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value
        """, (new_hash,))
    audit('CHANGE_PASSWORD', 'Admin password changed')
    notify('Password changed', 'The admin password was updated successfully.', 'info')
    return jsonify({'message': 'Password updated successfully'})

# ── SESSION PING ────────────────────────────────────────────────────────────────────────────────

@app.route('/session/ping', methods=['POST'])
@login_required
def session_ping():
    session.modified = True
    return jsonify({'ok': True, 'timeout_minutes': 30})

# ── GLOBAL ERROR HANDLERS ──────────────────────────────────────────────────────────────────────────────

@app.errorhandler(404)
def not_found(e):
    if request.accept_mimetypes.accept_json and not request.accept_mimetypes.accept_html:
        return jsonify({'error': 'Not found'}), 404
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(e):
    if request.accept_mimetypes.accept_json and not request.accept_mimetypes.accept_html:
        return jsonify({'error': 'Internal server error'}), 500
    return render_template('500.html'), 500

# ── ENTRYPOINT ────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    app.run(
        host='0.0.0.0',
        port=int(os.environ.get('PORT', 5000)),
        debug=os.environ.get('FLASK_ENV') != 'production',
    )